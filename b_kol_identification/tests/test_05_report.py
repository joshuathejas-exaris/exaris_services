import importlib.util, os
_S = os.path.join(os.path.dirname(__file__), "..", "05_generate_report.py")
_spec = importlib.util.spec_from_file_location("rep", _S)
mod = importlib.util.module_from_spec(_spec); _spec.loader.exec_module(mod)

DATA = {"indication":"Obesity","client_drug":"Ozempic","generated_at":"2026-07-09T10:00:00",
  "pca_terms":[{"term_key":"CF_OBESITY","term_en":"Obesity"}],
  "hcps":[{"s_customer_id":"10","name":"Anna Berg","city":"Berlin","specialty":"Innere Medizin",
           "rating":"A","verified_web_count":3,"verified_pubmed_count":4,"kol_score":0.7231800766,"latest_year":2024,
           "tier":"A","rising_star":False,"pub_by_year":{"2023":2,"2024":2},
           "theme_labels":[{"term_key":"CF_OBESITY","term_en":"Obesity","count":5}],
           "top_quotes":[{"quote":"patient improved","url":"http://x","sentiment":"positive"}]}],
  "coauthor_edges":[{"hcp_a":"10","hcp_b":"ext1","shared_pmids":2,"a_name":"Anna Berg","b_name":"Ext P","b_external":True}],
  "comention_edges":[]}


def test_stat_cards_show_kol_count_and_no_digiscore():
    html = mod.render_stat_cards(DATA)
    assert "1" in html
    assert "digi" not in html.lower()

def test_stat_cards_verified_sources_kpi_uses_raw_counts_not_composite():
    # kol_score is now a 0-1 composite; the "Verified sources" KPI must sum the raw
    # verified_web_count + verified_pubmed_count across HCPs, not kol_score (which
    # would previously sum to a near-meaningless fraction).
    html = mod.render_stat_cards(DATA)
    total = sum(h.get("verified_web_count", 0) + h.get("verified_pubmed_count", 0) for h in DATA["hcps"])
    assert f'<div class="v">{total}</div>' in html

def test_kol_table_shows_composite_score_formatted_and_header():
    html = mod.render_kol_table(DATA["hcps"], top_n=25)
    h = DATA["hcps"][0]
    assert "Anna Berg" in html
    assert "Obesity" in html                       # theme
    assert "<th>Composite score</th>" in html
    assert "<th>Verified sources</th>" not in html
    assert f'{h["kol_score"]:.2f}' in html         # formatted composite (0.72), not raw float
    assert f'({h["verified_web_count"]}w / {h["verified_pubmed_count"]}p)' in html

def test_network_lists_external_collaborator():
    html = mod.render_network(DATA["coauthor_edges"], DATA["comention_edges"], DATA["hcps"])
    assert "Ext P" in html

def test_build_report_html_is_selfcontained():
    import re
    html = mod.build_report_html(DATA)
    assert html.strip().startswith("<!DOCTYPE html>")
    assert "Anna Berg" in html
    # No external resources the browser would auto-fetch.
    assert "<link " not in html          # no stylesheet/font links
    assert 'src="http' not in html       # no external <script>/<img> sources
    assert "@import" not in html         # no CSS @import of remote sheets
    # No external href inside a <link> or <script> tag (a plain quote link is allowed).
    for tag in re.findall(r"<(?:link|script)\b[^>]*>", html):
        assert 'href="http' not in tag and 'src="http' not in tag


# ── Supplementary tests for the ported renderers (v2 field names) ──────────────

def test_rising_stars_only_lists_flagged_hcps():
    hcps = [dict(DATA["hcps"][0]),
            {**DATA["hcps"][0], "s_customer_id": "11", "name": "Boris Klein",
             "rising_star": True, "pub_by_year": {"2023": 0, "2024": 3}}]
    html = mod.render_rising_stars(hcps, ["2023", "2024"])
    assert "Boris Klein" in html
    assert "Anna Berg" not in html  # not flagged as rising

def test_rising_stars_empty_when_none_flagged():
    assert mod.render_rising_stars(DATA["hcps"], ["2023", "2024"]) == ""

def test_rising_stars_numbers_come_from_verified_pubmed_years_not_pub_by_year():
    # The Rising badge (Stage 04) is computed from verified_pubmed_years, so the card's
    # recent/prior/ratio text must be computed from that field too -- not from the
    # unverified/candidate pub_by_year -- or the displayed numbers won't justify the badge.
    h = {**DATA["hcps"][0], "rising_star": True,
         "pub_by_year": {"2020": 100, "2021": 100},
         "verified_pubmed_years": {"2024": 2, "2025": 1}}
    html = mod.render_rising_stars([h], ["2020", "2021", "2024", "2025"])
    assert "<b>3</b> recent" in html and "<b>0</b> prior" in html
    assert "100" not in html

def _rising_hcp_with_traj():
    return {"name": "Rita Stern", "specialty": "Innere Medizin", "city": "Kiel",
            "rising_star": True, "relevant_tenure": 2, "kol_score": 0.71,
            "verified_pubmed_years": {"2022": 2, "2023": 3},
            "total_pub_by_year": {"2022": 3, "2023": 4},
            "theme_labels": [{"term_key": "CF_OBESITY", "term_en": "Obesity", "count": 3}],
            "norm_relevance": 0.8, "norm_reach": 0.5, "norm_ratio": 0.6,
            "factor_contributions": {"relevance": 0.48, "reach": 0.12, "ratio": 0.09},
            "reach": {"distinct_coauthors": 5, "distinct_affiliations": 4},
            "ratio": {"ratio": 0.6, "denominator": 10},
            "verified_web_count": 2, "verified_pubmed_count": 5,
            "top_quotes": [{"quote": "q", "url": "http://x"}],
            "score_trajectory": [{"year": 2021, "score": 0.3, "tier": "C", "tenure": 0},
                                 {"year": 2022, "score": 0.5, "tier": "B", "tenure": 1},
                                 {"year": 2023, "score": 0.71, "tier": "B", "tenure": 2}]}


def test_rising_stars_has_score_development_section_with_chart():
    html = mod.render_rising_stars([_rising_hcp_with_traj()], ["2021", "2022", "2023"],
                                   weights={"relevance": 0.6, "reach": 0.25, "ratio": 0.15},
                                   t_a=0.8, t_b=0.4)
    assert "Score development" in html
    assert "<polyline" in html            # the dev line chart rendered


def test_rising_stars_row_has_score_breakdown():
    html = mod.render_rising_stars([_rising_hcp_with_traj()], ["2021", "2022", "2023"],
                                   weights={"relevance": 0.6, "reach": 0.25, "ratio": 0.15})
    assert "score-breakdown" in html      # the <details> drill-down is present in the table
    assert "how it was scored" in html


def test_thematic_heatmap_uses_theme_labels_not_cf_by_term():
    html = mod.render_thematic_heatmap(DATA["hcps"], DATA["pca_terms"], top_n=20)
    assert "Anna Berg" in html and "Obesity" in html and "5" in html

def test_thematic_heatmap_cell_is_populated_not_blank():
    # theme_labels[].term_key ("CF_OBESITY") must match pca_terms[].term_key so the
    # HCP's count actually lands in a heatmap cell instead of every lookup missing.
    html = mod.render_thematic_heatmap(DATA["hcps"], DATA["pca_terms"], top_n=20)
    assert "background:rgba(" in html      # a real (non-"transparent") cell was rendered
    assert ">5</td>" in html               # the count itself is visible in that cell
    assert 'background:transparent">5</td>' not in html

def test_regional_groups_by_city_with_tier_breakdown():
    hcps = DATA["hcps"] + [{**DATA["hcps"][0], "s_customer_id": "12", "name": "Carla Weiss",
                            "city": "Munich", "tier": "B"}]
    html = mod.render_regional(hcps)
    assert "Berlin" in html and "Munich" in html

def test_profiles_render_quotes_and_verified_source_breakdown():
    html = mod.render_profiles(DATA["hcps"], ["2023", "2024"], top_n=10)
    h = DATA["hcps"][0]
    # The actual verified quote text must render.
    assert h["top_quotes"][0]["quote"] in html
    # The composite score must render formatted and clearly labeled as a composite --
    # not mislabeled as a source count.
    assert f'Composite score {h["kol_score"]:.2f}' in html
    # The ACTUAL verified source total (web+pubmed count) must render separately from
    # the composite.
    verified_total = h["verified_web_count"] + h["verified_pubmed_count"]
    assert f'{verified_total} verified sources' in html
    assert f'{h["verified_web_count"]} web' in html
    assert f'{h["verified_pubmed_count"]} pubmed' in html

def test_no_composite_or_digi_fields_leak_into_full_report():
    html = mod.build_report_html(DATA)
    assert "composite_score" not in html
    assert "digi_score" not in html.lower()


def test_report_has_grouped_sidebar_nav():
    html = mod.build_report_html(DATA)
    assert 'class="sidebar"' in html
    for group in ("OVERVIEW", "ANALYSIS", "PROFILES"):
        assert group in html
    for item in ("Executive Dashboard", "KOL Ranking", "Rising Stars",
                 "Thematic Distribution", "Regional Distribution",
                 "Collaboration Network", "KOL Profiles"):
        assert item in html

def test_report_has_exactly_one_active_panel_and_tab_script():
    html = mod.build_report_html(DATA)
    assert html.count('class="panel active"') == 1   # first panel active on load
    assert "function showTab(" in html
    assert "js-tabs" in html

def test_tab_id_slugifies_label():
    assert mod.tab_id("Rising Stars") == "tab-rising-stars"


def test_year_axis_is_fixed_span_ending_at_anchor():
    data = {"anchor_year": 2023, "pub_history_years": 20, "hcps": []}
    axis = mod.build_year_axis(data)
    assert axis[0] == "2004" and axis[-1] == "2023"
    assert len(axis) == 20
    assert all(isinstance(y, str) for y in axis)

def test_year_axis_falls_back_to_present_years_without_anchor():
    data = {"hcps": [{"pub_by_year": {"2019": 1, "2021": 2}},
                     {"pub_by_year": {"2020": 1}}]}
    assert mod.build_year_axis(data) == ["2019", "2020", "2021"]

def test_report_uses_20y_axis_in_profile_label():
    data = {**DATA, "anchor_year": 2023, "pub_history_years": 20}
    html = mod.build_report_html(data)
    assert "2004" in html and "2023" in html   # full span rendered in a spark label


def test_write_excel_creates_one_row_per_kol(tmp_path):
    out = tmp_path / "k.xlsx"
    mod.write_excel(DATA, str(out))
    import openpyxl
    wb = openpyxl.load_workbook(out); ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "Name" in headers and "Verified sources" in headers and "Tier" in headers
    assert "Composite score" in headers
    assert ws.max_row == 1 + len(DATA["hcps"])
    # "Composite score" holds the 0-1 float; "Verified sources" holds the raw count.
    h = DATA["hcps"][0]
    comp_idx = headers.index("Composite score")
    vs_idx = headers.index("Verified sources")
    row2 = [c.value for c in ws[2]]
    assert row2[comp_idx] == h["kol_score"]
    assert row2[vs_idx] == h["verified_web_count"] + h["verified_pubmed_count"]


# ── Task 7: network graph, score drill-down, per-section explainers ────────────

def test_network_svg_is_selfcontained():
    svg = mod.render_network_svg(
        [{"a_name": "A B", "b_name": "C D", "shared_pmids": 2, "b_external": False}],
        [{"name": "A B", "reach": 3, "affiliation": "Uni A"}, {"name": "C D", "reach": 1, "affiliation": "Uni B"}])
    assert "<svg" in svg and "http://" not in svg and "https://" not in svg
    assert "Uni A" in svg   # affiliation surfaced (label/title)

def test_select_network_includes_externals_and_drops_isolated_kols():
    kol_nodes = [{"name": "Anna Berg", "reach": 4, "affiliation": "Uni A"},
                 {"name": "Carl Ott", "reach": 2, "affiliation": "Uni B"},
                 {"name": "Lonely Kol", "reach": 0, "affiliation": ""}]  # no edges -> dropped
    edges_in = [
        {"a_name": "Anna Berg", "b_name": "Carl Ott", "shared_pmids": 3, "b_external": False},
        {"a_name": "Anna Berg", "b_name": "Ext P", "shared_pmids": 2, "b_external": True},
    ]
    nodes, edges = mod._select_network(edges_in, kol_nodes)
    names = {n["name"] for n in nodes}
    assert "Anna Berg" in names and "Carl Ott" in names and "Ext P" in names
    assert "Lonely Kol" not in names                      # isolated KOL dropped
    assert any(n["name"] == "Ext P" and n["kol"] is False for n in nodes)
    assert len(edges) == 2

def test_force_layout_spreads_nodes_within_bounds_and_deterministic():
    names = [f"n{i}" for i in range(12)]
    edges = [{"a": "n0", "b": f"n{i}", "w": 1} for i in range(1, 6)]
    p1 = mod._force_layout(names, edges, 1080, 620)
    p2 = mod._force_layout(names, edges, 1080, 620)
    assert p1 == p2                                        # deterministic
    xs = [x for x, _ in p1.values()]; ys = [y for _, y in p1.values()]
    assert all(0 <= x <= 1080 for x in xs) and all(0 <= y <= 620 for y in ys)
    # genuinely 2-D spread, not all on one circle: both axes cover real range
    assert max(xs) - min(xs) > 200 and max(ys) - min(ys) > 150

def test_score_breakdown_shows_three_factors():
    hcp = {"name": "X", "kol_score": 0.72,
           "factor_contributions": {"relevance": 0.5, "reach": 0.15, "ratio": 0.07},
           "norm_relevance": 0.83, "norm_reach": 0.6, "norm_ratio": 0.47,
           "reach": {"distinct_coauthors": 6, "distinct_affiliations": 3},
           "ratio": {"ratio": 0.47, "denominator": 17},
           "top_quotes": [{"quote": "q", "url": "u", "sentiment": "positive"}]}
    html = mod.render_score_breakdown(hcp, {"relevance": 0.6, "reach": 0.25, "ratio": 0.15})
    for token in ("Relevance", "Reach", "Ratio", "0.72", "6", "q"):
        assert token in html

def test_as_of_banner_only_when_backtesting():
    assert mod.as_of_banner(2021, "2021") != ""
    assert mod.as_of_banner(2025, "latest") == ""


def test_network_node_prefers_real_affiliation_over_city():
    # Spec 9: the network graph should show real co-author affiliations on hover,
    # not the HCP's practice city (same city != same institution).
    data = {**DATA, "hcps": [{**DATA["hcps"][0], "affiliations": ["Uni Klinikum X"]}]}
    html = mod.build_report_html(data)
    assert "Uni Klinikum X" in html


def test_network_node_falls_back_to_city_without_affiliations():
    # Spec 9: when a KOL has no recorded co-author affiliations, the network
    # graph's hover title must fall back to their practice city instead of
    # leaving it blank.
    data = {**DATA, "hcps": [{**DATA["hcps"][0], "affiliations": []}]}
    html = mod.build_report_html(data)
    assert "Anna Berg — Berlin · reach 0</title>" in html
    assert 'data-aff="Berlin"' in html


# ── Task 11: stacked total-vs-relevant per-year bars ────────────────────────

def test_render_year_bars_stacks_total_and_relevant():
    svg = mod.render_year_bars({"2017": 4, "2018": 6}, {"2017": 1, "2018": 3},
                               ["2016", "2017", "2018"])
    assert svg.startswith("<svg") and svg.count("<rect") >= 4      # total + relevant per active year
    assert "</svg>" in svg

def test_render_year_bars_empty_when_no_data():
    assert mod.render_year_bars({}, {}, ["2016", "2017"]) == ""


def test_render_score_dev_chart_has_bands_and_line():
    traj = [{"year": 2016, "score": 0.1, "tier": "C", "tenure": 1},
            {"year": 2017, "score": 0.4, "tier": "B", "tenure": 2},
            {"year": 2018, "score": 0.9, "tier": "A", "tenure": 3}]
    svg = mod.render_score_dev_chart(traj, thresh_a=0.8, thresh_b=0.4)
    assert svg.startswith("<svg") and "<polyline" in svg
    assert svg.count("<rect") >= 3           # three tier bands
    assert "</svg>" in svg

def test_render_score_dev_chart_empty_for_short_series():
    assert mod.render_score_dev_chart([], 0.8, 0.4) == ""
    assert mod.render_score_dev_chart([{"year": 2018, "score": 0.5, "tier": "C", "tenure": 1}], 0.8, 0.4) == ""


def test_render_year_bars_has_axes_and_shared_width():
    svg = mod.render_year_bars({"2017": 4, "2018": 6}, {"2017": 1, "2018": 3},
                               ["2016", "2017", "2018"])
    assert svg.count("<line") >= 2                       # x-axis + y-axis
    assert f'width="{mod.PROFILE_CHART_W}"' in svg       # same width as the dev chart


# ── Task 13: report wiring — total pubs, career labels, disjoint counts ────────

def test_tenure_chip_shows_years_on_topic_or_empty():
    assert mod.tenure_chip({"relevant_tenure": 11}) == "11y on-topic"
    assert mod.tenure_chip({"relevant_tenure": None}) == ""   # web-only KOL, no pub tenure
    assert mod.tenure_chip({}) == ""


def test_kol_ranking_excludes_rising_stars_but_rising_tab_includes_them():
    # A high-scoring rising star (is_kol=False) must NOT appear in the KOL ranking table,
    # only in the Rising Stars tab. A real KOL (is_kol=True) must appear in the ranking.
    star = {**DATA["hcps"][0], "s_customer_id": "77", "name": "Rey Zoom",
            "kol_score": 0.99, "tier": None, "is_kol": False, "rising_star": True,
            "relevant_tenure": 2, "verified_pubmed_years": {"2024": 2, "2025": 1}}
    kol = {**DATA["hcps"][0], "s_customer_id": "10", "name": "Anna Berg",
           "is_kol": True, "rising_star": False, "tier": "A", "relevant_tenure": 9}
    html = mod.build_report_html({**DATA, "hcps": [star, kol]})
    # Split on the section heading (content-only); the nav uses the bare label "Rising Stars".
    ranking = html.split("<h2>Rising Stars</h2>")[0]   # nav + overview/KOL-ranking panel
    assert "Anna Berg" in ranking                # the KOL is in the ranking
    assert "Rey Zoom" not in ranking             # the rising star is NOT in the ranking
    assert "Rey Zoom" in html                    # but appears elsewhere (Rising Stars tab)


# ── Final-review fix: tier=None (non-KOL/rising-star) must not crash rendering ──

def test_report_survives_none_tier_top_of_list_hcp():
    # Stage 04's assign_tiers sets tier=None for every non-KOL HCP (rising stars and
    # floor-failing HCPs). A high-scoring non-KOL near the top of the score-sorted
    # list must not crash the report with `None.lower()` / a KeyError.
    breakout_star = {**DATA["hcps"][0], "s_customer_id": "99", "name": "Nora Fast",
                     "kol_score": 0.95, "tier": None, "is_kol": False, "rising_star": True,
                     "breakout": True, "relevant_tenure": 2}
    hcps = [breakout_star] + DATA["hcps"]
    data = {**DATA, "hcps": hcps}
    html = mod.build_report_html(data)   # must not raise
    assert "Nora Fast" in html


def test_kol_table_and_profiles_render_none_tier_without_raising():
    breakout_star = {**DATA["hcps"][0], "s_customer_id": "99", "name": "Nora Fast",
                     "kol_score": 0.95, "tier": None, "is_kol": False, "rising_star": True}
    hcps = [breakout_star] + DATA["hcps"]
    table_html = mod.render_kol_table(hcps, top_n=25)   # must not raise
    assert "Nora Fast" in table_html
    assert 'class="pill none"' in table_html
    profiles_html = mod.render_profiles(hcps, ["2023", "2024"], top_n=10)   # must not raise
    assert "Nora Fast" in profiles_html


def test_established_new_to_topic_detects_veteran_pivot():
    # publishes since 2008 (long total span) but first RELEVANT year 2017 (short tenure)
    hcp = {"total_pub_by_year": {"2008": 2, "2012": 3, "2017": 1, "2018": 2},
           "relevant_tenure": 2}
    assert mod.established_new_to_topic(hcp) is True
    assert mod.established_new_to_topic({"total_pub_by_year": {"2017": 1}, "relevant_tenure": 2}) is False


def test_profiles_omit_score_breakdown_dropdown():
    html = mod.render_profiles(DATA["hcps"], ["2023", "2024"], top_n=10)
    assert "score-breakdown" not in html
    assert "how it was scored" not in html


def test_profiles_omit_tenure_sticker():
    h = dict(DATA["hcps"][0]); h["relevant_tenure"] = 7
    html = mod.render_profiles([h], ["2023", "2024"], top_n=10)
    assert 'pill stage' not in html


# ── Task 5: Excel "LLM Wiki Verdicts" sheet ─────────────────────────────────────

def test_wiki_verdict_rows_mark_counted_and_rejected():
    hcps = [{"s_customer_id": "10", "name": "Anna Berg"}]
    sources = {"hcps": [{"s_customer_id": "10",
                         "web_sources": [{"source_id": "w1", "kind": "web", "url": "http://a"},
                                         {"source_id": "w2", "kind": "web", "url": "http://b"}],
                         "pubmed_sources": [{"source_id": "111", "kind": "pubmed",
                                             "pmid": "111", "url": "http://pm/111",
                                             "full_text": "SECRET BODY"}]}]}
    wiki = {"hcps": [{"s_customer_id": "10",
                      "claims": [{"source_id": "w1", "statement": "s1", "themes": ["Obesity"],
                                  "sentiment": "positive", "verified": True},
                                 {"source_id": "111", "statement": "s2", "themes": ["NASH"],
                                  "sentiment": "neutral", "verified": True}]}]}
    rows = mod.build_wiki_verdict_rows(hcps, sources, wiki)
    hdr = mod.WIKI_VERDICT_HEADERS
    verdicts = {r[hdr.index("URL")]: r[hdr.index("Verdict")] for r in rows}
    assert verdicts["http://a"] == "counted"      # w1 produced a claim
    assert verdicts["http://b"] == "rejected"      # w2 handed over, no claim
    assert verdicts["http://pm/111"] == "counted"
    assert all("SECRET BODY" not in str(c) for r in rows for c in r)   # full_text never leaks


def test_write_excel_gates_verdict_rows_on_both_sources_and_wiki_loaded(tmp_path):
    """Regression test: if sources.json loads but wiki.json is missing, show the
    'not available' note instead of misleading 'rejected' verdicts for every source."""
    import json, openpyxl
    out = tmp_path / "k.xlsx"
    # Write a real minimal sources.json with one HCP and one web source
    sources_path = tmp_path / "sources.json"
    sources_json = {
        "hcps": [{
            "s_customer_id": "10",
            "web_sources": [{"source_id": "w1", "kind": "web", "url": "http://example.com"}],
            "pubmed_sources": []
        }]
    }
    with open(sources_path, "w", encoding="utf-8") as f:
        json.dump(sources_json, f)

    # Call write_excel with sources present but wiki missing (None path)
    mod.write_excel(DATA, str(out), sources_path=str(sources_path), wiki_path=None)

    # Load and inspect the Excel file
    wb = openpyxl.load_workbook(out)
    ws2 = wb["LLM Wiki Verdicts"]

    # Should contain only the header row + one note row (no verdict rows)
    assert ws2.max_row == 2, f"Expected 2 rows (header + note), got {ws2.max_row}"

    # Check that the note text is present in row 2, column B
    note_cell = ws2["B2"].value
    assert "not available" in note_cell, f"Expected 'not available' in note, got: {note_cell}"

    # No verdict column should contain "rejected" (that would be misleading)
    verdict_col_idx = mod.WIKI_VERDICT_HEADERS.index("Verdict") + 1  # 1-indexed
    for row_idx in range(2, ws2.max_row + 1):
        cell = ws2.cell(row=row_idx, column=verdict_col_idx)
        assert cell.value != "rejected", f"Row {row_idx} has misleading 'rejected' verdict when wiki was missing"


def test_score_year_rows_one_per_trajectory_year():
    hcps = [{"name": "Kai", "score_trajectory": [
                {"year": 2022, "score": 0.5, "relevance": 10, "reach": 3, "ratio": 0.6, "tenure": 1, "tier": "B"},
                {"year": 2023, "score": 0.7, "relevance": 12, "reach": 5, "ratio": 0.65, "tenure": 2, "tier": "A"}]},
            {"name": "NoTraj", "score_trajectory": []}]
    rows = mod.build_score_year_rows(hcps)
    hdr = mod.SCORE_YEAR_HEADERS
    assert len(rows) == 2                                  # NoTraj contributes nothing
    assert [r[hdr.index("Year")] for r in rows] == [2022, 2023]
    assert rows[1][hdr.index("Tier")] == "A"
    assert rows[0][hdr.index("Relevance")] == 10
