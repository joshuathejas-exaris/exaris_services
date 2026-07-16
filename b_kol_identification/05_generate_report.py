"""
Stage 05: Generate the KOL report (HTML top-25) + Excel.
Reads:  data/kol_final.json
Writes: results/kol_report_<ts>.html  and  results/kol_report_<ts>.xlsx
"""
import configparser, html as _html, json, logging, math, os, re, sys
from datetime import datetime

logging.basicConfig(format="%(asctime)s [%(levelname)s] %(message)s", level=logging.INFO)
log = logging.getLogger(__name__)
_DIR = os.path.dirname(__file__)

PALETTE = {
    "ink": "#1b2430", "muted": "#5c6774", "line": "#e2e7ee", "bg": "#f4f6f8", "card": "#fff",
    "accent": "#2f4a7c", "teal": "#0d7d74", "violet": "#6d5ac0", "amber": "#b7791f", "emerald": "#1f8a5b",
    "tierA": "#1f8a5b", "tierB": "#3b5b92", "tierC": "#6b7684",
    "pos": "#0d7d74", "neu": "#6b7684", "neg": "#b4432f",
}


def _esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;").replace("'", "&#39;"))


def _rgba(hex_color, alpha):
    h = hex_color.lstrip("#")
    r, g, b = (int(h[i:i + 2], 16) for i in (0, 2, 4))
    return f"rgba({r},{g},{b},{alpha})"


# Fallback composite weights (Task 6 defaults) used when the caller does not pass
# the run's actual [scoring] weights through to the report renderer.
DEFAULT_WEIGHTS = {"relevance": 0.60, "reach": 0.25, "ratio": 0.15}

RISING_MAX_TENURE_DEFAULT = 3

# Shared width for the two per-profile charts so their x-axes span the same length and
# the publication bars sit directly above the matching points on the score line below.
PROFILE_CHART_W = 320

# Fallback tier percentiles (Stage 04 config defaults) used when the caller does not
# pass the run's actual [scoring] tier_a_percentile/tier_b_percentile through.
DEFAULT_TIER_PCTS = (85.0, 60.0)


def tenure_chip(hcp):
    """Concrete on-topic tenure for display, e.g. '11y on-topic' — the number of years
    since the HCP's first indication-relevant publication. Empty when unknown (web-only
    KOLs with no publication tenure)."""
    t = hcp.get("relevant_tenure")
    return f"{t}y on-topic" if isinstance(t, int) and t > 0 else ""


def established_new_to_topic(hcp, min_total_span=8, max_relevant_tenure=3):
    """Long overall publication history but only recently relevant to THIS indication."""
    yrs = [int(y) for y in (hcp.get("total_pub_by_year") or {}).keys() if str(y).isdigit()]
    if not yrs:
        return False
    total_span = max(yrs) - min(yrs) + 1
    ten = hcp.get("relevant_tenure")
    return total_span >= min_total_span and ten is not None and ten <= max_relevant_tenure


def _kol_tier_thresholds(hcps, a_pct, b_pct):
    """Composite-score A/B thresholds over the KOL pool only (is_kol==True) —
    mirrors 04_assemble_kols.py::assign_tiers's threshold math so the score-development
    chart's tier bands line up with an HCP's actual assigned tier."""
    kol_scores = sorted(h.get("kol_score", 0) for h in hcps if h.get("is_kol"))
    if not kol_scores:
        return float("inf"), float("inf")
    n = len(kol_scores)
    t_a = kol_scores[min(int(n * a_pct / 100), n - 1)]
    t_b = kol_scores[min(int(n * b_pct / 100), n - 1)]
    return t_a, t_b


def as_of_banner(anchor_year, as_of_year_cfg) -> str:
    if not as_of_year_cfg or str(as_of_year_cfg).strip().lower() == "latest":
        return ""
    return (f'<div class="asof-banner">Backtest view — PubMed capped at {anchor_year}. '
            f'Web sources are timestamp-free and shown as-is (frozen across years).</div>')


def section_explainer(text: str) -> str:
    return f'<p class="explainer"><strong>How to read this:</strong> {_html.escape(text)}</p>'


def render_score_breakdown(hcp: dict, weights: dict) -> str:
    fc = hcp.get("factor_contributions", {})
    reach = hcp.get("reach", {}); ratio = hcp.get("ratio", {})
    rows = [
        ("Relevance", weights["relevance"], hcp.get("norm_relevance", 0), fc.get("relevance", 0),
         f'{hcp.get("verified_web_count",0)+hcp.get("verified_pubmed_count",0)} verified sources'),
        ("Reach", weights["reach"], hcp.get("norm_reach", 0), fc.get("reach", 0),
         f'{reach.get("distinct_coauthors",0)} co-authors, {reach.get("distinct_affiliations",0)} institutions'),
        ("Ratio", weights["ratio"], hcp.get("norm_ratio", 0), fc.get("ratio", 0),
         f'{ratio.get("ratio",0):.0%} of {ratio.get("denominator",0)} total sources'),
    ]
    tr = "".join(
        f'<tr><td>{name}</td><td>{w:.2f}</td><td>{norm:.2f}</td><td>{contrib:.3f}</td><td>{_html.escape(ev)}</td></tr>'
        for (name, w, norm, contrib, ev) in rows)
    quotes = "".join(
        f'<li>“{_html.escape(q.get("quote", ""))}” '
        f'<a href="{_html.escape(q.get("url",""))}">source</a></li>'
        for q in hcp.get("top_quotes", []))
    return (f'<details class="score-breakdown"><summary>Composite {hcp.get("kol_score",0):.2f} — how it was scored</summary>'
            f'<table><thead><tr><th>Factor</th><th>Weight</th><th>Norm</th><th>Contribution</th><th>Evidence</th></tr></thead>'
            f'<tbody>{tr}</tbody></table>'
            f'<div class="score-quotes"><strong>Evidence quotes:</strong><ul>{quotes}</ul></div></details>')


def _select_network(coauthor_edges: list, kol_nodes: list, max_external: int = 40):
    """Pick the nodes+edges to draw: all KOLs + the most-connected external co-authors,
    keeping only edges whose endpoints are both in the set, then drop isolated nodes.

    Returns (nodes, edges) where each node is
    {name, reach, affiliation, kol: bool, degree: int} and each edge is
    {a, b, w, external: bool} (w = shared PubMed papers)."""
    kol_by_name = {n["name"]: n for n in kol_nodes if n.get("name")}
    # rank external co-authors by how many KOLs they bridge, then shared-paper volume
    ext = {}
    for e in coauthor_edges:
        a, b = e.get("a_name"), e.get("b_name")
        if a and b and e.get("b_external") and b not in kol_by_name:
            s = ext.setdefault(b, {"deg": set(), "shared": 0})
            s["deg"].add(a)
            s["shared"] += int(e.get("shared_pmids", 0) or 0)
    ranked = sorted(ext.items(), key=lambda kv: (len(kv[1]["deg"]), kv[1]["shared"], kv[0]), reverse=True)
    keep_ext = {name for name, _ in ranked[:max_external]}

    edges = []
    for e in coauthor_edges:
        a, b = e.get("a_name"), e.get("b_name")
        if not a or not b or a == b or a not in kol_by_name:
            continue
        if b in kol_by_name or b in keep_ext:
            edges.append({"a": a, "b": b, "w": int(e.get("shared_pmids", 0) or 0),
                          "external": bool(e.get("b_external") and b not in kol_by_name)})

    degree = {}
    for e in edges:
        degree[e["a"]] = degree.get(e["a"], 0) + 1
        degree[e["b"]] = degree.get(e["b"], 0) + 1

    nodes = []
    for name in degree:  # only connected nodes — no lonely circles on a ring
        kn = kol_by_name.get(name)
        nodes.append({"name": name,
                      "reach": (kn or {}).get("reach", 0),
                      "affiliation": (kn or {}).get("affiliation", ""),
                      "kol": name in kol_by_name,
                      "degree": degree[name]})
    return nodes, edges


def _force_layout(names: list, edges: list, width: int, height: int, iterations: int = 90) -> dict:
    """Deterministic Fruchterman-Reingold layout → {name: (x, y)} within a margin.

    Deterministic (golden-angle seed, no RNG) so a given graph always lays out the
    same way — reproducible across runs and testable."""
    n = len(names)
    if n == 0:
        return {}
    margin = 60
    W, H = width - 2 * margin, height - 2 * margin
    if n == 1:
        return {names[0]: (margin + W / 2, margin + H / 2)}
    ga = math.pi * (3 - math.sqrt(5))  # golden angle → even initial spread
    pos = {}
    for i, nm in enumerate(names):
        rad = 0.45 * math.sqrt((i + 0.5) / n)
        pos[nm] = [0.5 * W + rad * W * math.cos(i * ga), 0.5 * H + rad * H * math.sin(i * ga)]
    adj = {}
    for e in edges:
        key = (e["a"], e["b"])
        adj[key] = adj.get(key, 0) + e.get("w", 1)
    k = 0.9 * math.sqrt((W * H) / n)  # ideal edge length
    t = W * 0.10
    cool = t / (iterations + 1)
    for _ in range(iterations):
        disp = {nm: [0.0, 0.0] for nm in names}
        for i in range(n):
            ni = names[i]
            for j in range(i + 1, n):
                nj = names[j]
                dx, dy = pos[ni][0] - pos[nj][0], pos[ni][1] - pos[nj][1]
                dist = math.hypot(dx, dy) or 0.01
                f = (k * k) / dist
                ux, uy = dx / dist, dy / dist
                disp[ni][0] += ux * f; disp[ni][1] += uy * f
                disp[nj][0] -= ux * f; disp[nj][1] -= uy * f
        for (a, b), w in adj.items():
            if a not in pos or b not in pos:
                continue
            dx, dy = pos[a][0] - pos[b][0], pos[a][1] - pos[b][1]
            dist = math.hypot(dx, dy) or 0.01
            f = (dist * dist) / k * (1 + 0.25 * math.log(1 + w))
            ux, uy = dx / dist, dy / dist
            disp[a][0] -= ux * f; disp[a][1] -= uy * f
            disp[b][0] += ux * f; disp[b][1] += uy * f
        for nm in names:
            dx, dy = disp[nm]
            d = math.hypot(dx, dy) or 0.01
            pos[nm][0] = min(W, max(0.0, pos[nm][0] + (dx / d) * min(d, t)))
            pos[nm][1] = min(H, max(0.0, pos[nm][1] + (dy / d) * min(d, t)))
        t = max(t - cool, W * 0.005)
    return {nm: (margin + pos[nm][0], margin + pos[nm][1]) for nm in names}


def render_network_svg(coauthor_edges: list, kol_nodes: list, width: int = 1080, height: int = 620) -> str:
    """Force-directed co-authorship graph as self-contained SVG. KOLs + their most
    connected external co-authors are nodes (sized by reach); edges are shared PubMed
    authorship (width = shared papers, dashed = external), labelled with the count.
    Node <title>/data-aff carry affiliations; edge <title> carries the relation."""
    nodes, edges = _select_network(coauthor_edges, kol_nodes)
    if not nodes:
        return ('<svg class="net-graph" width="100%" height="120" viewBox="0 0 600 120">'
                '<text x="300" y="62" text-anchor="middle" fill="#5c6774" font-size="14">'
                'No co-authorship connections among the top KOLs.</text></svg>')
    names = [n["name"] for n in nodes]
    pos = _force_layout(names, edges, width, height)
    maxreach = max((n["reach"] for n in nodes), default=1) or 1
    maxw = max((e["w"] for e in edges), default=1) or 1
    label_cut = max(2, maxw // 3)

    edge_svg = []
    for e in edges:
        (x1, y1), (x2, y2) = pos[e["a"]], pos[e["b"]]
        sw = 1.2 + 3.8 * (e["w"] / maxw)
        dash = ' stroke-dasharray="5 4"' if e["external"] else ""
        rel = f'{e["w"]} shared PubMed paper' + ("s" if e["w"] != 1 else "")
        edge_svg.append(
            f'<line class="net-edge" data-a="{_esc(e["a"])}" data-b="{_esc(e["b"])}" '
            f'x1="{x1:.1f}" y1="{y1:.1f}" x2="{x2:.1f}" y2="{y2:.1f}" '
            f'stroke="#b6c2d1" stroke-width="{sw:.1f}"{dash}>'
            f'<title>{_esc(e["a"])} ↔ {_esc(e["b"])} — {rel}</title></line>')
        if e["w"] >= label_cut:
            edge_svg.append(
                f'<text class="net-edge-label" data-a="{_esc(e["a"])}" data-b="{_esc(e["b"])}" '
                f'x="{(x1 + x2) / 2:.1f}" y="{(y1 + y2) / 2:.1f}">{e["w"]}</text>')

    node_svg = []
    for n in nodes:
        x, y = pos[n["name"]]
        if n["kol"]:
            rad = 8 + 16 * (n["reach"] / maxreach)
            fill = PALETTE["accent"]
        else:
            rad = 5 + min(n["degree"], 5)
            fill = PALETTE["violet"]
        aff = str(n.get("affiliation", "") or ("" if n["kol"] else "external co-author"))
        title = _esc(n["name"]) + (f' — {_esc(aff)}' if aff else "") + \
            (f' · reach {n["reach"]}' if n["kol"] else "")
        g = (f'<g class="net-node{"" if n["kol"] else " ext"}" data-name="{_esc(n["name"])}" '
             f'data-aff="{_esc(aff)}" data-kol="{"1" if n["kol"] else "0"}" data-reach="{n["reach"]}" '
             f'data-x="{x:.1f}" data-y="{y:.1f}" transform="translate({x:.1f},{y:.1f})">'
             f'<circle r="{rad:.1f}" fill="{fill}" stroke="#fff" stroke-width="1.5">'
             f'<title>{title}</title></circle>')
        if n["kol"] or n["degree"] >= 2:  # label KOLs + bridge externals; hide singleton externals
            g += f'<text class="net-label" x="0" y="{-rad - 4:.1f}" text-anchor="middle">{_esc(n["name"])}</text>'
        node_svg.append(g + "</g>")

    return (f'<svg class="net-graph" viewBox="0 0 {width} {height}" '
            f'preserveAspectRatio="xMidYMid meet" width="100%" height="{height}">'
            f'<g class="net-zoom">{"".join(edge_svg)}{"".join(node_svg)}</g></svg>')


def _recent_prior(pub_by_year):
    """Recent (current + prior year) vs. prior publication counts, from a year->count dict."""
    years = {int(y): int(c) for y, c in (pub_by_year or {}).items() if str(y).isdigit()}
    if not years:
        return 0, 0
    cur = max(years)
    recent = sum(c for y, c in years.items() if y >= cur - 1)
    prior = sum(c for y, c in years.items() if y < cur - 1)
    return recent, prior


def render_stat_cards(data):
    hcps = data["hcps"]
    tiers = {t: sum(1 for h in hcps if h.get("tier") == t) for t in "ABC"}
    # KOLs and Rising stars are disjoint buckets (Stage 04: is_kol requires NOT rising_star),
    # so these two counts must never double-count the same HCP.
    kols = sum(1 for h in hcps if h.get("is_kol"))
    rising = sum(1 for h in hcps if h.get("rising_star"))
    total_sources = sum(h.get("verified_web_count", 0) + h.get("verified_pubmed_count", 0) for h in hcps)
    cards = [("KOLs", kols), ("Tier A", tiers["A"]), ("Tier B", tiers["B"]),
             ("Tier C", tiers["C"]), ("Rising stars", rising), ("Verified sources", total_sources)]
    cells = "".join(
        f'<div class="stat"><div class="v">{v}</div><div class="k">{_esc(k)}</div></div>' for k, v in cards)
    return f'<div class="stats">{cells}</div>'


def render_kol_table(hcps, top_n, weights=None):
    weights = weights or DEFAULT_WEIGHTS
    rows = ""
    for i, h in enumerate(hcps[:top_n], 1):
        themes = ", ".join(_esc(t["term_en"]) for t in h.get("theme_labels", [])[:3])
        badge = f'<span class="pill {(h.get("tier") or "none").lower()}">{h.get("tier") or "—"}</span>'
        tc = tenure_chip(h)
        stage = f' <span class="pill stage">{_esc(tc)}</span>' if tc else ""
        rows += (f'<tr><td>{i}</td><td>{badge}{stage}</td>'
                 f'<td><b>{_esc(h["name"])}</b><br><span class="muted">{_esc(h["specialty"])}</span></td>'
                 f'<td>{_esc(h["city"])}</td>'
                 f'<td><b>{h["kol_score"]:.2f}</b> '
                 f'<span class="muted">({h.get("verified_web_count",0)}w / {h.get("verified_pubmed_count",0)}p)</span>'
                 f'{render_score_breakdown(h, weights)}</td>'
                 f'<td>{h.get("total_pubmed_sources", 0)}</td>'
                 f'<td>{themes}</td></tr>')
    return (f'<table><thead><tr><th>#</th><th>Tier</th><th>Name / Specialty</th><th>City</th>'
            f'<th>Composite score</th><th>Total pubs</th><th>Themes</th></tr></thead><tbody>{rows}</tbody></table>')


def render_sparkline(pub_by_year, all_years, width=80, height=24):
    counts = [pub_by_year.get(y, 0) for y in all_years]
    max_v = max(counts, default=0) or 1
    n = max(len(all_years), 1)
    bw = width / n - 1
    bars = []
    for i, c in enumerate(counts):
        bh = max(2.0, c / max_v * (height - 4))
        x = i * (width / n)
        y = height - bh
        bars.append(f'<rect x="{x:.1f}" y="{y:.1f}" width="{bw:.1f}" '
                    f'height="{bh:.1f}" fill="{PALETTE["accent"]}"/>')
    return f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}">{"".join(bars)}</svg>'


def render_year_bars(total_by_year, relevant_by_year, all_years, width=PROFILE_CHART_W, height=54):
    """Grouped per-year bars: light column = all publications that year, dark inner
    column = the verified-relevant subset. Drawn on an x/y axis frame, and sharing the
    score-development chart's width + horizontal insets so the columns line up above the
    score line. Inline SVG (no CDN)."""
    tot = {str(y): int(v) for y, v in (total_by_year or {}).items()}
    rel = {str(y): int(v) for y, v in (relevant_by_year or {}).items()}
    if not tot and not rel:
        return ""
    years = list(all_years)
    peak = max([tot.get(y, 0) for y in years] + [rel.get(y, 0) for y in years] + [1])
    # Same horizontal insets as render_score_dev_chart (pad_l/pad_r = 6) so both charts'
    # plotted regions start and end at the same x.
    pad_l, pad_r, pad_t, pad_b = 6, 6, 6, 12
    plot_w = width - pad_l - pad_r
    base = height - pad_b
    plot_h = base - pad_t
    n = max(len(years), 1)
    bw = plot_w / n
    pad = bw * 0.2
    rects, labels = [], []
    for i, y in enumerate(years):
        x = pad_l + i * bw + pad
        w = bw - 2 * pad
        th = (tot.get(y, 0) / peak) * plot_h
        rh = (rel.get(y, 0) / peak) * plot_h
        if tot.get(y, 0):
            rects.append(f'<rect x="{x:.1f}" y="{base - th:.1f}" width="{w:.1f}" '
                         f'height="{th:.1f}" fill="{PALETTE["line"]}"/>')
        if rel.get(y, 0):
            rects.append(f'<rect x="{x:.1f}" y="{base - rh:.1f}" width="{w:.1f}" '
                         f'height="{rh:.1f}" fill="{PALETTE["accent"]}"/>')
        if y.endswith("0") or y.endswith("5"):
            labels.append(f'<text x="{x + w/2:.1f}" y="{height - 1}" font-size="7" '
                          f'text-anchor="middle" fill="{PALETTE["muted"]}">{y[2:]}</text>')
    axes = (f'<line x1="{pad_l}" y1="{pad_t}" x2="{pad_l}" y2="{base:.1f}" '
            f'stroke="{PALETTE["muted"]}" stroke-width="1"/>'
            f'<line x1="{pad_l}" y1="{base:.1f}" x2="{width - pad_r}" y2="{base:.1f}" '
            f'stroke="{PALETTE["muted"]}" stroke-width="1"/>')
    return (f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" '
            f'role="img" aria-label="publications per year, total vs relevant">'
            f'{axes}{"".join(rects)}{"".join(labels)}</svg>')


def render_score_dev_chart(trajectory, thresh_a, thresh_b, width=PROFILE_CHART_W, height=120, rising_max=RISING_MAX_TENURE_DEFAULT):
    """Line chart of composite score over years with A/B/C tier bands and a marker at
    the year the HCP crossed from rising-star tenure into KOL tenure. Inline SVG."""
    pts = [p for p in (trajectory or []) if isinstance(p.get("score"), (int, float))]
    if len(pts) < 2:
        return ""
    pad_l, pad_r, pad_t, pad_b = 6, 6, 6, 14
    plot_w = width - pad_l - pad_r
    plot_h = height - pad_t - pad_b
    ta = max(0.0, min(1.0, float(thresh_a))) if thresh_a != float("inf") else 1.0
    tb = max(0.0, min(1.0, float(thresh_b))) if thresh_b != float("inf") else 0.6
    def sy(v):  # score 0..1 -> y
        return pad_t + (1 - max(0.0, min(1.0, v))) * plot_h
    bands = [  # (top_v, bottom_v, colour)
        (1.0, ta, PALETTE.get("tierA", "#1f8a5b")),
        (ta, tb, PALETTE.get("tierB", "#3b5b92")),
        (tb, 0.0, PALETTE.get("tierC", "#6b7684")),
    ]
    rects = "".join(
        f'<rect x="{pad_l}" y="{sy(top):.1f}" width="{plot_w}" '
        f'height="{max(0.0, sy(bot) - sy(top)):.1f}" fill="{col}" opacity="0.10"/>'
        for top, bot, col in bands)
    n = len(pts)
    xs = [pad_l + (i / (n - 1)) * plot_w for i in range(n)]
    poly = " ".join(f"{xs[i]:.1f},{sy(pts[i]['score']):.1f}" for i in range(n))
    line = f'<polyline points="{poly}" fill="none" stroke="{PALETTE["accent"]}" stroke-width="2"/>'
    dots = "".join(f'<circle cx="{xs[i]:.1f}" cy="{sy(pts[i]["score"]):.1f}" r="2.2" '
                   f'fill="{PALETTE["accent"]}"/>' for i in range(n))
    # tenure-crossing marker: first year tenure exceeds the rising limit
    marker = ""
    for i, p in enumerate(pts):
        if p.get("tenure", 0) == rising_max + 1:
            marker = (f'<line x1="{xs[i]:.1f}" y1="{pad_t}" x2="{xs[i]:.1f}" '
                      f'y2="{pad_t + plot_h}" stroke="{PALETTE.get("amber", "#b7791f")}" '
                      f'stroke-width="1" stroke-dasharray="3 2"/>'
                      f'<text x="{xs[i]:.1f}" y="{height - 3}" font-size="7" text-anchor="middle" '
                      f'fill="{PALETTE.get("amber", "#b7791f")}">→ KOL tenure</text>')
            break
    return (f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" '
            f'role="img" aria-label="score development over years">'
            f'{rects}{marker}{line}{dots}</svg>')


def render_rising_stars(hcps, all_years, weights=None, t_a=float("inf"), t_b=float("inf"),
                        rising_max=RISING_MAX_TENURE_DEFAULT):
    weights = weights or DEFAULT_WEIGHTS
    stars = [h for h in hcps if h.get("rising_star")]
    if not stars:
        return ""
    # Table (on top) — one row per rising star with score, tenure and momentum.
    trows = ""
    for i, h in enumerate(stars, 1):
        # The Rising badge (Stage 04) is computed from verified_pubmed_years, so the
        # displayed recent/prior/ratio must come from the same (verified) field -- not
        # the unverified/candidate pub_by_year -- or the numbers won't justify the badge.
        recent, prior = _recent_prior(h.get("verified_pubmed_years", {}))
        ratio = f"{recent / max(prior, 1):.1f}×" if prior > 0 else "New voice"
        breakout = ' <span class="pill breakout">Breakout</span>' if h.get("breakout") else ""
        tenure = h.get("relevant_tenure")
        tenure_txt = f"{tenure}y on-topic" if isinstance(tenure, int) and tenure > 0 else "—"
        themes = ", ".join(_esc(t["term_en"]) for t in h.get("theme_labels", [])[:3])
        trows += (f'<tr><td>{i}</td>'
                  f'<td><b>{_esc(h.get("name",""))}</b> <span class="pill rise">Rising</span>{breakout}<br>'
                  f'<span class="muted">{_esc(h.get("specialty",""))}</span></td>'
                  f'<td>{_esc(h.get("city",""))}</td><td>{tenure_txt}</td>'
                  f'<td><b>{h.get("kol_score",0):.2f}</b>{render_score_breakdown(h, weights)}</td>'
                  f'<td><b>{recent}</b> recent vs <b>{prior}</b> prior &middot; {ratio}</td>'
                  f'<td>{themes}</td></tr>')
    table = (f'<table><thead><tr><th>#</th><th>Name / Specialty</th><th>City</th><th>Tenure</th>'
             f'<th>Composite score</th><th>Recent vs prior (verified pubs)</th><th>Themes</th>'
             f'</tr></thead><tbody>{trows}</tbody></table>')
    # Publication bars — total vs indication-relevant per year, same style as KOL profiles.
    bar_cards = ""
    for h in stars:
        bars = render_year_bars(h.get("total_pub_by_year", {}), h.get("verified_pubmed_years", {}), all_years)
        if not bars:
            continue
        bar_cards += (f'<div class="rising-card"><b>{_esc(h.get("name",""))}</b>'
                      f'<div style="margin:.4rem 0">{bars}'
                      f'<span class="muted spark-label">pubs/yr — total vs relevant</span></div></div>')
    bars_block = (f'<h3>Publication trajectory — total vs indication-relevant</h3>'
                  f'<div class="rising-grid">{bar_cards}</div>') if bar_cards else ""
    # Score development (separate section) — composite score over years with tier bands,
    # the same chart used on KOL profile cards.
    dev_cards = ""
    for h in stars:
        chart = render_score_dev_chart(h.get("score_trajectory", []), t_a, t_b, rising_max=rising_max)
        if not chart:
            continue
        dev_cards += (f'<div class="rising-card"><b>{_esc(h.get("name",""))}</b>'
                      f'<div style="margin:.4rem 0">{chart}'
                      f'<span class="muted spark-label">score development</span></div></div>')
    dev_block = (f'<h3>Score development — composite score over time</h3>'
                 f'<div class="rising-grid">{dev_cards}</div>') if dev_cards else ""
    return f'<h2>Rising Stars</h2>{table}{bars_block}{dev_block}'


def render_established_new_callout(hcps):
    """Callout for veteran researchers (long overall publication history) whose
    verified engagement with THIS indication only recently began."""
    matches = [h for h in hcps if established_new_to_topic(h)]
    if not matches:
        return ""
    items = ""
    for h in matches:
        yrs = [int(y) for y in (h.get("total_pub_by_year") or {}).keys() if str(y).isdigit()]
        since = min(yrs) if yrs else "?"
        items += (f'<li><b>{_esc(h.get("name",""))}</b> — publishing since {since}, '
                  f'only {h.get("relevant_tenure","?")}y relevant to this indication</li>')
    return (f'<div class="callout"><h3>Established, new to this indication</h3>'
            f'<p class="muted">Veteran researchers with a long overall publication history whose '
            f'verified engagement with this specific indication only recently began.</p>'
            f'<ul>{items}</ul></div>')


def render_thematic_heatmap(hcps, pca_terms, top_n=20):
    top = hcps[:top_n]
    keys = [t["term_key"] for t in pca_terms]

    def _tc(h):
        return {t["term_key"]: t["count"] for t in h.get("theme_labels", [])}

    col_max = {k: max((_tc(h).get(k, 0) for h in top), default=1) or 1 for k in keys}
    headers = "".join(f'<th title="{_esc(t["term_en"])}">{_esc(t["term_en"])}</th>' for t in pca_terms)
    rows = ""
    for h in top:
        tc = _tc(h)
        badge = f'<span class="pill {(h.get("tier") or "none").lower()}">{h.get("tier") or "—"}</span>'
        cells = ""
        for k in keys:
            count = tc.get(k, 0)
            alpha = round(count / col_max[k] * 0.75, 2) if count else 0
            bg = _rgba(PALETTE["accent"], alpha) if count else "transparent"
            cells += f'<td style="text-align:center;background:{bg}">{count or ""}</td>'
        rows += f'<tr><td class="muted">{_esc(h.get("name",""))}</td><td>{badge}</td>{cells}</tr>'
    return (f'<h2>Thematic Distribution — Top {top_n}</h2>'
            f'<div class="hmap-wrap"><table><thead><tr><th>KOL</th><th>Tier</th>{headers}</tr></thead>'
            f'<tbody>{rows}</tbody></table></div>')


def render_regional(hcps, top_n=20):
    from collections import defaultdict
    city_data = defaultdict(lambda: {"A": 0, "B": 0, "C": 0})
    for h in hcps:
        city = h.get("city") or "Unknown"
        tier = h.get("tier")
        if tier in ("A", "B", "C"):
            city_data[city][tier] += 1
        else:
            city_data[city]  # still register the city even if untiered (non-KOL) HCPs live there
    ranked = sorted(city_data.items(), key=lambda x: -(x[1]["A"] + x[1]["B"] + x[1]["C"]))[:top_n]
    if not ranked:
        return ""
    max_total = max(d["A"] + d["B"] + d["C"] for _, d in ranked) or 1
    bar_max_px = 260
    rows = ""
    for city, d in ranked:
        total = d["A"] + d["B"] + d["C"]
        scale = total / max_total
        wa = int(d["A"] / total * bar_max_px * scale) if total else 0
        wb = int(d["B"] / total * bar_max_px * scale) if total else 0
        wc = int(d["C"] / total * bar_max_px * scale) if total else 0
        segs = ""
        if wa:
            segs += f'<div style="width:{wa}px;background:{PALETTE["tierA"]}"></div>'
        if wb:
            segs += f'<div style="width:{wb}px;background:{PALETTE["tierB"]}"></div>'
        if wc:
            segs += f'<div style="width:{wc}px;background:{PALETTE["tierC"]}"></div>'
        labels = ""
        if d["A"]:
            labels += f'<span class="pill a">{d["A"]}A</span> '
        if d["B"]:
            labels += f'<span class="pill b">{d["B"]}B</span> '
        if d["C"]:
            labels += f'<span class="pill c">{d["C"]}C</span>'
        rows += (f'<div class="city-row"><div class="city-label" title="{_esc(city)}">{_esc(city)}</div>'
                 f'<div class="city-bar" style="width:{bar_max_px}px">{segs}</div>'
                 f'<div class="city-count muted">{total} &nbsp; {labels}</div></div>')
    return f'<h2>Regional Distribution</h2>{rows}'


def render_network(coauthor_edges, comention_edges, hcps):
    def rows(edges, kind):
        out = ""
        for e in sorted(edges, key=lambda x: x.get("shared_pmids", x.get("count", 0)), reverse=True):
            if kind == "co":
                tag = ' <span class="pill ext">external</span>' if e.get("b_external") else ""
                out += (f'<tr><td>{_esc(e["a_name"])}</td><td>{_esc(e["b_name"])}{tag}</td>'
                        f'<td>{e["shared_pmids"]} shared PMIDs</td></tr>')
            else:
                out += (f'<tr><td>{_esc(e["from_name"])}</td><td>{_esc(e["to_name"])}</td>'
                        f'<td>{e["count"]} web mentions</td></tr>')
        return out or '<tr><td colspan="3" class="muted">none</td></tr>'
    return (f'<h3>PubMed co-authorship</h3><table><tbody>{rows(coauthor_edges,"co")}</tbody></table>'
            f'<h3>Web co-mentions</h3><table><tbody>{rows(comention_edges,"men")}</tbody></table>')


_SENT_COLOR_KEY = {"positive": "pos", "negative": "neg"}


def render_profiles(hcps, all_years, top_n=10, weights=None, tier_thresholds=None, rising_max=RISING_MAX_TENURE_DEFAULT):
    weights = weights or DEFAULT_WEIGHTS
    t_a, t_b = tier_thresholds or (float("inf"), float("inf"))
    year_range = f"{all_years[0]}–{all_years[-1]}" if all_years else ""
    cards = ""
    for h in hcps[:top_n]:
        tier = h.get("tier") or "—"
        badge = f'<span class="pill {(h.get("tier") or "none").lower()}">{tier}</span>'
        year_bars = render_year_bars(h.get("total_pub_by_year", {}), h.get("verified_pubmed_years", {}), all_years)
        dev_chart = render_score_dev_chart(h.get("score_trajectory", []), t_a, t_b, rising_max=rising_max)
        themes = "".join(f'<span class="tag">{_esc(t["term_en"])}</span>' for t in h.get("theme_labels", []))
        verified_total = h.get("verified_web_count", 0) + h.get("verified_pubmed_count", 0)
        meta = (f'<div class="muted">Composite score {h.get("kol_score",0):.2f} &middot; '
                f'{verified_total} verified sources '
                f'({h.get("verified_web_count",0)} web / {h.get("verified_pubmed_count",0)} pubmed) '
                f'&middot; {h.get("total_pubmed_sources",0)} total publications</div>')
        quotes = ""
        for q in h.get("top_quotes", [])[:3]:
            color = PALETTE[_SENT_COLOR_KEY.get(q.get("sentiment"), "neu")]
            link = (f' <a href="{_esc(q["url"])}">source</a>' if q.get("url") else "")
            quotes += (f'<div class="quote" style="border-left:3px solid {color}">'
                       f'“{_esc(q.get("quote",""))}”{link}</div>')
        charts = (f'<div style="margin:.5rem 0">{year_bars}'
                  f'<span class="muted spark-label">pubs/yr, total vs relevant ({year_range})</span></div>')
        if dev_chart:
            charts += (f'<div style="margin:.5rem 0">{dev_chart}'
                       f'<span class="muted spark-label">score development</span></div>')
        cards += (
            f'<div class="profile-card">'
            f'<div style="display:flex;justify-content:space-between">'
            f'<div><b>{_esc(h.get("name",""))}</b><br>'
            f'<span class="muted">{_esc(h.get("specialty",""))} · {_esc(h.get("city",""))}</span></div>'
            f'<div>{badge}</div></div>'
            f'{meta}'
            f'{charts}'
            f'<div>{themes}</div>{quotes}</div>'
        )
    return f'<h2>Individual KOL Profiles — Top {top_n}</h2><div class="profile-grid">{cards}</div>'


def build_year_axis(data):
    """Fixed pub_history_years-length axis of string years ending at anchor_year.
    Falls back to the union of years present in pub_by_year for pre-anchor JSON."""
    anchor = data.get("anchor_year")
    span = int(data.get("pub_history_years") or 20)
    if anchor:
        anchor = int(anchor)
        return [str(y) for y in range(anchor - span + 1, anchor + 1)]
    present = sorted({y for h in data.get("hcps", []) for y in h.get("pub_by_year", {})})
    return [str(y) for y in present]


def tab_id(label: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", label.lower()).strip("-")
    return "tab-" + (slug or "x")


TAB_SCRIPT = """
<script>
function showTab(id){
  var ps = document.querySelectorAll('.panel');
  for (var i = 0; i < ps.length; i++){ ps[i].classList.toggle('active', ps[i].id === id); }
  var ns = document.querySelectorAll('.nav-item');
  for (var j = 0; j < ns.length; j++){
    var on = ns[j].getAttribute('href') === '#' + id;
    ns[j].classList.toggle('active', on);
    if (on) { ns[j].setAttribute('aria-current', 'page'); }
    else { ns[j].removeAttribute('aria-current'); }
  }
  return false;
}
document.addEventListener('DOMContentLoaded', function(){
  document.body.classList.add('js-tabs');
  var f = document.querySelector('.panel');
  if (f) { showTab(f.id); }
});
</script>
"""


NETWORK_SCRIPT = """
<script>
(function(){
  var svg = document.querySelector('svg.net-graph');
  if (!svg || !svg.querySelector('.net-zoom')) return;
  var zoom = svg.querySelector('.net-zoom');
  var wrap = svg.closest('.network-svg-wrap');
  var tip = wrap ? wrap.querySelector('.net-tooltip') : null;
  var vb = svg.viewBox.baseVal;
  var scale = 1, tx = 0, ty = 0;
  function apply(){ zoom.setAttribute('transform', 'translate(' + tx + ',' + ty + ') scale(' + scale + ')'); }
  var panning = false, sx = 0, sy = 0;
  svg.addEventListener('mousedown', function(ev){ panning = true; sx = ev.clientX; sy = ev.clientY; svg.style.cursor = 'grabbing'; });
  window.addEventListener('mousemove', function(ev){ if (!panning) return; tx += ev.clientX - sx; ty += ev.clientY - sy; sx = ev.clientX; sy = ev.clientY; apply(); });
  window.addEventListener('mouseup', function(){ panning = false; svg.style.cursor = ''; });
  svg.addEventListener('wheel', function(ev){
    ev.preventDefault();
    var r = svg.getBoundingClientRect();
    var mx = (ev.clientX - r.left) / r.width * vb.width;
    var my = (ev.clientY - r.top) / r.height * vb.height;
    var f = ev.deltaY < 0 ? 1.12 : 0.89;
    var ns = Math.min(5, Math.max(0.25, scale * f));
    tx = mx - (mx - tx) * (ns / scale); ty = my - (my - ty) * (ns / scale); scale = ns; apply();
  }, { passive: false });
  function showTip(html, ev){
    if (!tip || !wrap) return;
    tip.innerHTML = html; tip.style.display = 'block';
    var r = wrap.getBoundingClientRect();
    tip.style.left = (ev.clientX - r.left + 14) + 'px';
    tip.style.top = (ev.clientY - r.top + 14) + 'px';
  }
  function hideTip(){ if (tip) tip.style.display = 'none'; }
  var nodes = svg.querySelectorAll('.net-node');
  for (var i = 0; i < nodes.length; i++){
    (function(g){
      g.addEventListener('mousemove', function(ev){
        var kol = g.getAttribute('data-kol') === '1';
        var aff = g.getAttribute('data-aff');
        var html = '<b>' + g.getAttribute('data-name') + '</b>' + (kol ? '<span class="tip-tag">KOL</span>' : '<span class="tip-tag ext">co-author</span>');
        if (kol) html += '<br>reach: ' + g.getAttribute('data-reach') + ' co-authors';
        if (aff) html += '<br>' + aff;
        showTip(html, ev);
      });
      g.addEventListener('mouseleave', hideTip);
    })(nodes[i]);
  }
  var edges = svg.querySelectorAll('.net-edge');
  for (var j = 0; j < edges.length; j++){
    (function(l){
      l.addEventListener('mousemove', function(ev){
        var t = l.querySelector('title');
        showTip(t ? t.textContent : '', ev);
      });
      l.addEventListener('mouseleave', hideTip);
    })(edges[j]);
  }
})();
</script>
"""


def _render_sidebar(groups):
    """groups: list of (group_label, [(item_label, panel_html), ...]). Sticky left nav
    beside a content pane of panels; first item active on load; degrades to full scroll
    without JS. Empty groups/items are skipped."""
    nav = ['<nav class="sidebar" role="tablist" aria-label="Report sections">']
    panels = []
    first = True
    for group_label, items in groups:
        items = [it for it in items if it and it[1]]
        if not items:
            continue
        nav.append(f'<div class="nav-group-label">{_esc(group_label)}</div>')
        for item_label, body in items:
            tid = tab_id(item_label)
            active = " active" if first else ""
            current = ' aria-current="page"' if first else ""
            nav.append(f'<a class="nav-item{active}" role="tab" id="{tid}-btn" '
                       f'href="#{tid}" aria-controls="{tid}"{current} '
                       f'onclick="return showTab(\'{tid}\')">{_esc(item_label)}</a>')
            panels.append(f'<section class="panel{active}" role="tabpanel" id="{tid}" '
                          f'aria-labelledby="{tid}-btn">\n{body}\n</section>')
            first = False
    nav.append("</nav>")
    content = '<main class="content">\n' + "\n".join(panels) + "\n</main>"
    return '<div class="layout">\n' + "\n".join(nav) + "\n" + content + "\n</div>"


def build_report_html(data, weights=None, as_of_year_cfg="latest", tier_pcts=None, rising_max=None):
    weights = weights or DEFAULT_WEIGHTS
    tier_pcts = tier_pcts or DEFAULT_TIER_PCTS
    rising_max = RISING_MAX_TENURE_DEFAULT if rising_max is None else rising_max
    all_years = build_year_axis(data)
    top_n = 25
    css = f"""
      body{{margin:0;background:{PALETTE['bg']};color:{PALETTE['ink']};
        font:15px/1.55 -apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif}}
      .wrap{{max-width:1800px;margin:0 auto;padding:28px 12px 64px}}
      h1{{margin:4px 0}} h2{{border-top:1px solid {PALETTE['line']};padding-top:14px;margin-top:34px}}
      table{{border-collapse:collapse;width:100%;font-size:13px;margin:10px 0}}
      th,td{{border:1px solid {PALETTE['line']};padding:7px 10px;text-align:left;vertical-align:top}}
      th{{background:#eef2f7}} .muted{{color:{PALETTE['muted']}}}
      .stats{{display:grid;grid-template-columns:repeat(6,1fr);gap:12px;margin:14px 0}}
      .stat{{background:{PALETTE['card']};border:1px solid {PALETTE['line']};border-radius:10px;padding:14px;text-align:center}}
      .stat .v{{font-size:24px;font-weight:800;color:{PALETTE['accent']}}} .stat .k{{font-size:12px;color:{PALETTE['muted']}}}
      .pill{{font-size:11px;font-weight:700;padding:1px 7px;border-radius:20px}}
      .pill.a{{background:#e7f5ee;color:{PALETTE['tierA']}}} .pill.b{{background:#eaf0f9;color:{PALETTE['tierB']}}}
      .pill.c{{background:#eef1f5;color:{PALETTE['tierC']}}} .pill.rise{{background:#fbf1dd;color:{PALETTE['amber']}}}
      .pill.ext{{background:#efecfa;color:{PALETTE['violet']}}}
      .pill.stage{{background:#eef2f7;color:{PALETTE['muted']}}}
      .pill.breakout{{background:#e7f5ee;color:{PALETTE['emerald']}}}
      .callout{{background:#fdf3e3;border:1px solid {PALETTE['amber']};border-radius:8px;
        padding:10px 14px;margin:14px 0}}
      .callout h3{{margin:0 0 6px;font-size:14px;color:{PALETTE['ink']}}}
      .callout ul{{margin:6px 0 0;padding-left:18px;font-size:13px}}
      .tag{{display:inline-block;background:#eef2f7;color:{PALETTE['teal']};font-size:11px;
        padding:1px 7px;border-radius:10px;margin:2px 3px 0 0}}
      .rising-grid,.profile-grid{{display:grid;gap:12px;margin:14px 0}}
      .rising-grid{{grid-template-columns:repeat(auto-fill,minmax(220px,1fr))}}
      .profile-grid{{grid-template-columns:repeat(auto-fill,minmax(300px,1fr))}}
      .rising-card,.profile-card{{background:{PALETTE['card']};border:1px solid {PALETTE['line']};
        border-radius:10px;padding:12px 14px}}
      .rising-card{{border-top:3px solid {PALETTE['amber']}}}
      .rising-card svg,.profile-card svg{{max-width:100%;height:auto;display:block}}
      .spark-label{{font-size:11px;margin-left:6px}}
      .hmap-wrap{{overflow-x:auto}}
      .hmap-wrap th{{writing-mode:vertical-rl;transform:rotate(180deg);font-size:11px;
        padding:6px 4px;min-width:30px}}
      .city-row{{display:flex;align-items:center;gap:10px;margin:5px 0;font-size:13px}}
      .city-label{{width:130px;text-align:right;white-space:nowrap;overflow:hidden;
        text-overflow:ellipsis;flex-shrink:0}}
      .city-bar{{height:16px;display:flex;border-radius:3px;overflow:hidden;flex-shrink:0}}
      .quote{{font-size:12px;font-style:italic;padding:4px 8px;margin-top:6px;color:{PALETTE['ink']}}}
      .quote a{{font-style:normal;font-size:11px;color:{PALETTE['accent']}}}
      .asof-banner{{background:#fdf3e3;border:1px solid {PALETTE['amber']};color:{PALETTE['ink']};
        border-radius:8px;padding:10px 14px;margin:10px 0;font-size:13px}}
      .explainer{{background:#eef2f7;border-left:3px solid {PALETTE['accent']};color:{PALETTE['muted']};
        border-radius:0 6px 6px 0;padding:6px 12px;margin:8px 0 14px;font-size:13px}}
      .explainer strong{{color:{PALETTE['ink']}}}
      details.score-breakdown{{margin-top:6px;font-size:12px}}
      details.score-breakdown summary{{cursor:pointer;color:{PALETTE['accent']};font-weight:600}}
      details.score-breakdown table{{font-size:12px;margin:6px 0}}
      .score-quotes{{margin-top:6px}} .score-quotes ul{{margin:4px 0;padding-left:18px}}
      .network-svg-wrap{{position:relative;overflow:hidden;margin:10px 0;
        border:1px solid {PALETTE['line']};border-radius:10px;background:{PALETTE['card']}}}
      svg.net-graph{{display:block;cursor:grab;touch-action:none;
        background:radial-gradient(circle at 1px 1px,#eef2f7 1px,transparent 0) 0 0/22px 22px}}
      .net-edge{{stroke:#b6c2d1}} .net-edge:hover{{stroke:{PALETTE['accent']}}}
      .net-edge-label{{fill:{PALETTE['muted']};font-size:10px;text-anchor:middle;pointer-events:none;
        paint-order:stroke;stroke:{PALETTE['card']};stroke-width:3px}}
      .net-node{{cursor:pointer}}
      .net-node:hover circle{{stroke:{PALETTE['ink']};stroke-width:2.5}}
      .net-label{{font-size:11px;fill:{PALETTE['ink']};pointer-events:none;font-weight:600;
        paint-order:stroke;stroke:{PALETTE['card']};stroke-width:3px}}
      .net-tooltip{{position:absolute;display:none;pointer-events:none;z-index:5;max-width:260px;
        background:{PALETTE['ink']};color:#fff;font-size:12px;line-height:1.45;padding:7px 10px;
        border-radius:7px;box-shadow:0 4px 14px rgba(0,0,0,.25)}}
      .net-tooltip .tip-tag{{font-size:10px;background:{PALETTE['accent']};padding:1px 6px;
        border-radius:8px;margin-left:4px}}
      .net-tooltip .tip-tag.ext{{background:{PALETTE['violet']}}}
      .net-legend{{display:flex;gap:18px;flex-wrap:wrap;font-size:12px;color:{PALETTE['muted']};padding:8px 4px 0}}
      .net-legend .dot{{display:inline-block;width:11px;height:11px;border-radius:50%;
        margin-right:5px;vertical-align:middle}}
      .net-hint{{font-size:11px;color:{PALETTE['muted']};padding:2px 4px 6px}}
      .layout{{display:flex;gap:28px;align-items:flex-start;margin:18px 0 8px}}
      .sidebar{{flex:0 0 210px;position:sticky;top:16px;align-self:flex-start}}
      .content{{flex:1 1 auto;min-width:0}}
      .nav-group-label{{text-transform:uppercase;letter-spacing:.6px;font-size:11px;
        font-weight:700;color:{PALETTE['muted']};margin:16px 0 6px}}
      .nav-group-label:first-child{{margin-top:0}}
      .nav-item{{display:block;padding:6px 10px;margin:2px 0;border-radius:6px;
        color:{PALETTE['ink']};font-size:14px;text-decoration:none;border-left:3px solid transparent}}
      .nav-item:hover{{background:#eef2f7;color:{PALETTE['accent']}}}
      .nav-item.active{{background:#eef4fb;color:{PALETTE['accent']};font-weight:600;
        border-left-color:{PALETTE['accent']}}}
      body.js-tabs .panel{{display:none}} body.js-tabs .panel.active{{display:block}}
      .content h2:first-child{{margin-top:0;border-top:none;padding-top:0}}
      @media(max-width:720px){{.stats{{grid-template-columns:repeat(2,1fr)}}
        .layout{{flex-direction:column;gap:8px}} .sidebar{{position:static;flex-basis:auto;width:100%}}}}
    """
    top = data["hcps"]
    # KOL-only view for the KOL sections: rising stars live exclusively in the Rising
    # Stars tab, not the KOL ranking/profiles/heatmap/network. Fall back to the full
    # list for old-schema data (no is_kol field) so the report is never empty.
    kols = [h for h in top if h.get("is_kol")] or top
    # Stage 04 now threads through the HCP's actual co-author affiliation strings
    # (top_affiliations) -- real institutions, which is what conveys cross-org
    # activity on hover. Practice city is only a last-resort fallback for KOLs
    # with no PubMed co-author affiliation data at all.
    network_nodes = [{"name": h.get("name", ""),
                       "reach": h.get("reach", {}).get("distinct_coauthors", 0),
                       "affiliation": ", ".join(h.get("affiliations", [])) or h.get("city", "")} for h in kols]
    banner = as_of_banner(data.get("anchor_year"), as_of_year_cfg)
    # Score-dev chart tier bands: thresholds computed over the FULL final pool (not just
    # the top-N sliced for the report), mirroring Stage 04's assign_tiers exactly.
    t_a, t_b = _kol_tier_thresholds(kols, tier_pcts[0], tier_pcts[1])

    def _splice_explainer(section_html: str, explainer_text: str) -> str:
        """Insert section_explainer(...) right after a section's leading <h2>...</h2>."""
        exp = section_explainer(explainer_text)
        marker = "</h2>"
        idx = section_html.find(marker)
        if idx == -1:
            return exp + section_html
        idx += len(marker)
        return section_html[:idx] + exp + section_html[idx:]

    kol_ranking_section = _splice_explainer(
        f'<h2>KOL Ranking — Top {top_n}</h2>{render_kol_table(kols, top_n, weights)}',
        "KOLs are ranked by a composite score that blends Relevance (LLM-verified topical "
        "engagement), Reach (co-author and institution breadth), and Ratio (the share of the "
        "KOL's total output that is on-topic); expand a row's score to see each factor's "
        "weight and contribution. The 'Ny on-topic' tag is the KOL's tenure — the number of "
        "years since their first indication-relevant publication. Rising stars are a separate, "
        "mutually-exclusive bucket and are shown only in the Rising Stars tab, never here.")
    rising_section = _splice_explainer(
        render_rising_stars(top, all_years, weights=weights, t_a=t_a, t_b=t_b, rising_max=rising_max)
        or '<h2>Rising Stars</h2><p class="muted">No rising stars identified.</p>',
        "Rising stars are a separate bucket from KOLs — climbers, not yet arrived. An HCP is a "
        "rising star when their indication-relevant publication tenure is short (first on-topic "
        "paper within the last few years) and they are genuinely active (enough recent verified "
        "output). Because the buckets are mutually exclusive, no one here appears in the KOL "
        "ranking. A 'Breakout' tag marks a rising star whose composite score already reaches "
        "KOL Tier-A level. The bars below show each star's total vs. indication-relevant "
        "publications per year.")
    thematic_section = _splice_explainer(
        render_thematic_heatmap(kols, data.get("pca_terms", []), top_n=top_n),
        "Cell shading shows how concentrated a KOL's verified claims are on that theme "
        "relative to other KOLs in the top list -- darker cells mean more verified claims "
        "on that theme.")
    net_legend = (
        f'<div class="net-legend">'
        f'<span><span class="dot" style="background:{PALETTE["accent"]}"></span>KOL (size = co-author reach)</span>'
        f'<span><span class="dot" style="background:{PALETTE["violet"]}"></span>External co-author</span>'
        f'<span>line = shared PubMed papers (label = count · dashed = external)</span></div>'
        f'<div class="net-hint">Scroll to zoom · drag the background to pan · '
        f'hover a node for its institutions, or an edge for the shared-paper count.</div>')
    network_section = _splice_explainer(
        f'<h2>Collaboration network</h2>'
        f'<div class="network-svg-wrap">{render_network_svg(data.get("coauthor_edges", []), network_nodes)}'
        f'<div class="net-tooltip"></div></div>'
        f'{net_legend}'
        f'{render_network(data["coauthor_edges"], data["comention_edges"], kols)}',
        "This is the co-authorship graph: each node is a KOL (blue, sized by how many distinct "
        "co-authors they have) or a frequently-shared external co-author (purple); each line is "
        "shared PubMed authorship, thicker and labelled with the number of shared papers, dashed "
        "when the co-author is external. Only connected KOLs appear here. The tables below list every edge.")
    # OVERVIEW is a single panel: the stat cards sat alone in a one-row dashboard with
    # the rest of the space empty, so the KOL ranking is folded in beneath them.
    overview_panel = (f'<h2>Executive dashboard</h2>{render_stat_cards(data)}'
                      f'{render_established_new_callout(kols)}'
                      f'{kol_ranking_section}')
    profiles_section = _splice_explainer(
        render_profiles(kols, all_years, top_n=top_n, weights=weights, tier_thresholds=(t_a, t_b), rising_max=rising_max),
        "Each card's 'Ny on-topic' tag is the KOL's tenure — years since their first "
        "indication-relevant publication. In the score-development chart, the green / blue / grey "
        "background bands are the Tier A / B / C score ranges, so you can watch the line climb "
        "through the tiers over time; the dashed marker is the year the KOL crossed out of "
        "rising-star tenure. The chart replays each KOL's trajectory against a fixed yardstick — "
        "today's final pool of KOLs is the ruler for every year shown, so the line reflects the "
        "individual's own growth, not pool churn. Web sources are a constant baseline with no "
        "timestamps, so they contribute the same amount to every year. Today's LLM verification "
        "verdicts are applied back onto each historical year's PubMed record. This chart cannot "
        "show when an HCP entered or exited the KOL pool, or was demoted — that comparison is "
        "what the two-run backtest (Stage 06) is for.")
    groups = [
        ("OVERVIEW", [
            ("Executive Dashboard", overview_panel),
        ]),
        ("ANALYSIS", [
            ("Rising Stars", rising_section),
            ("Thematic Distribution", thematic_section),
            ("Regional Distribution", render_regional(kols)),
            ("Collaboration Network", network_section),
        ]),
        ("PROFILES", [
            ("KOL Profiles", profiles_section),
        ]),
    ]
    header = (f'<h1>KOL Identification — {_esc(data["indication"])}</h1>'
              f'<p class="muted">Client drug: {_esc(data["client_drug"])} · '
              f'generated {_esc(data["generated_at"])}</p>' + banner)
    body = header + "\n" + _render_sidebar(groups) + "\n" + TAB_SCRIPT + NETWORK_SCRIPT
    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>KOL Report — {_esc(data['indication'])}</title><style>{css}</style></head>
<body><div class="wrap">
{body}
</div></body></html>"""


WIKI_VERDICT_HEADERS = ["Rank", "Name", "Kind", "URL", "PMID", "Verdict",
                        "Verified claims", "Statements", "Themes", "Sentiments"]
_CELL_MAX = 1500  # keep joined statement cells readable, well under Excel's 32767 limit


def _load_json_safe(path):
    if not path:
        return None
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError):
        return None


def build_wiki_verdict_rows(hcps, sources_data, wiki_data):
    """One row per source handed to the LLM (from sources.json), joined to the verified
    claims it produced (from wiki.json), keyed by source_id. Verdict is 'counted' when the
    source yielded >=1 verified claim, else 'rejected'. Source full_text is never emitted."""
    src_by_id = {h.get("s_customer_id"): h for h in (sources_data or {}).get("hcps", [])}
    claims_by_hcp = {}
    for wh in (wiki_data or {}).get("hcps", []):
        by_src = {}
        for c in wh.get("claims", []):
            by_src.setdefault(str(c.get("source_id")), []).append(c)
        claims_by_hcp[wh.get("s_customer_id")] = by_src
    rows = []
    for rank, h in enumerate(hcps, 1):
        cid = h.get("s_customer_id")
        sh = src_by_id.get(cid)
        if not sh:
            continue
        by_src = claims_by_hcp.get(cid, {})
        for s in (sh.get("web_sources", []) + sh.get("pubmed_sources", [])):
            claims = by_src.get(str(s.get("source_id")), [])
            statements = " | ".join(c.get("statement", "") for c in claims)[:_CELL_MAX]
            themes = ", ".join(sorted({t for c in claims for t in (c.get("themes") or [])}))
            sentiments = ", ".join(sorted({c.get("sentiment", "") for c in claims if c.get("sentiment")}))
            rows.append([rank, h.get("name", ""), s.get("kind", ""), s.get("url", ""),
                         s.get("pmid", ""), "counted" if claims else "rejected",
                         len(claims), statements, themes, sentiments])
    return rows


SCORE_YEAR_HEADERS = ["Rank", "Name", "Year", "Composite score", "Relevance",
                      "Reach", "Ratio", "Tenure", "Tier"]


def build_score_year_rows(hcps):
    """One row per (HCP, trajectory year) from score_trajectory. HCPs with no trajectory
    contribute no rows. Mirrors the report's score-development chart data."""
    rows = []
    for rank, h in enumerate(hcps, 1):
        for p in (h.get("score_trajectory") or []):
            rows.append([rank, h.get("name", ""), p.get("year"),
                         round(float(p.get("score", 0)), 4), p.get("relevance"),
                         p.get("reach"), round(float(p.get("ratio", 0)), 4),
                         p.get("tenure"), p.get("tier")])
    return rows


def _autosize(ws, headers, max_w=60):
    from openpyxl.utils import get_column_letter
    for ci in range(1, len(headers) + 1):
        col = get_column_letter(ci)
        best = max((len(str(c.value)) for c in ws[col] if c.value is not None), default=0)
        ws.column_dimensions[col].width = min(max_w, best) + 2


def write_excel(data: dict, path: str, sources_path: str = None, wiki_path: str = None) -> None:
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "KOLs"
    headers = ["Rank", "Name", "Specialty", "City", "Tier", "Rising star",
               "Composite score", "Verified sources", "Web", "PubMed", "Latest year", "Top themes",
               "Representative quote", "Source URL",
               "norm_relevance", "norm_reach", "norm_ratio",
               "contribution_relevance", "contribution_reach", "contribution_ratio",
               "distinct_coauthors", "distinct_affiliations", "relevance_ratio",
               "total_publications", "relevant_tenure", "is_kol", "breakout"]
    ws.append(headers)
    for i, h in enumerate(data["hcps"], 1):
        q = (h.get("top_quotes") or [{}])[0]
        fc = h.get("factor_contributions", {})
        reach = h.get("reach", {}); ratio = h.get("ratio", {})
        verified_total = h.get("verified_web_count", 0) + h.get("verified_pubmed_count", 0)
        ws.append([i, h["name"], h["specialty"], h["city"], h.get("tier", ""),
                   "yes" if h.get("rising_star") else "",
                   h.get("kol_score", 0), verified_total,
                   h.get("verified_web_count", 0), h.get("verified_pubmed_count", 0),
                   h.get("latest_year", ""),
                   ", ".join(t["term_en"] for t in h.get("theme_labels", [])[:5]),
                   q.get("quote", ""), q.get("url", ""),
                   h.get("norm_relevance", 0), h.get("norm_reach", 0), h.get("norm_ratio", 0),
                   fc.get("relevance", 0), fc.get("reach", 0), fc.get("ratio", 0),
                   reach.get("distinct_coauthors", 0), reach.get("distinct_affiliations", 0),
                   ratio.get("ratio", 0),
                   h.get("total_pubmed_sources", 0), h.get("relevant_tenure", ""),
                   "yes" if h.get("is_kol") else "", "yes" if h.get("breakout") else ""])
    ws.freeze_panes = "A2"
    _autosize(ws, headers)

    # Sheet 2 — LLM Wiki Verdicts: one row per source handed to the LLM, with the
    # 'counted'/'rejected' verdict and the verified claim(s) it produced.
    ws2 = wb.create_sheet("LLM Wiki Verdicts")
    ws2.append(WIKI_VERDICT_HEADERS)
    src_data = _load_json_safe(sources_path)
    wiki_data = _load_json_safe(wiki_path)
    if src_data is None or wiki_data is None:
        note = ["", "sources.json / wiki.json not available — run stages 02–03"]
        ws2.append(note + [""] * (len(WIKI_VERDICT_HEADERS) - len(note)))
    else:
        for r in build_wiki_verdict_rows(data["hcps"], src_data, wiki_data):
            ws2.append(r)
    ws2.freeze_panes = "A2"
    _autosize(ws2, WIKI_VERDICT_HEADERS)

    # Sheet 3 — Score by Year: composite reconstruction per year (score-dev chart data).
    ws3 = wb.create_sheet("Score by Year")
    ws3.append(SCORE_YEAR_HEADERS)
    for r in build_score_year_rows(data["hcps"]):
        ws3.append(r)
    ws3.freeze_panes = "A2"
    _autosize(ws3, SCORE_YEAR_HEADERS)

    wb.save(path)


def main():
    import argparse
    p = argparse.ArgumentParser(); p.add_argument("--force", action="store_true")
    args = p.parse_args()
    cfg = configparser.ConfigParser(); cfg.read(os.path.join(_DIR, "config.ini"))
    sc = cfg["scoring"] if cfg.has_section("scoring") else {}
    weights = {"relevance": float(sc.get("weight_relevance", DEFAULT_WEIGHTS["relevance"])),
               "reach": float(sc.get("weight_reach", DEFAULT_WEIGHTS["reach"])),
               "ratio": float(sc.get("weight_ratio", DEFAULT_WEIGHTS["ratio"]))}
    tier_pcts = (float(sc.get("tier_a_percentile", DEFAULT_TIER_PCTS[0])),
                 float(sc.get("tier_b_percentile", DEFAULT_TIER_PCTS[1])))
    rising_max = int(sc.get("rising_star_max_tenure_years", RISING_MAX_TENURE_DEFAULT))
    as_of_year_cfg = cfg["funnel"].get("as_of_year", "latest") if cfg.has_section("funnel") else "latest"
    with open(os.path.join(_DIR, "data", "kol_final.json"), encoding="utf-8") as f:
        data = json.load(f)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    html_path = os.path.join(_DIR, "results", f"kol_report_{ts}.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(build_report_html(data, weights=weights, as_of_year_cfg=as_of_year_cfg, tier_pcts=tier_pcts, rising_max=rising_max))
    log.info(f"Wrote {html_path}")
    xlsx_path = os.path.join(_DIR, "results", f"kol_report_{ts}.xlsx")
    write_excel(data, xlsx_path,
                sources_path=os.path.join(_DIR, "data", "sources.json"),
                wiki_path=os.path.join(_DIR, "data", "wiki.json"))
    log.info(f"Wrote {xlsx_path}")

if __name__ == "__main__":
    main()
