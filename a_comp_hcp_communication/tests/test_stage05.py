import os

from conftest import load_stage

mod = load_stage("05_generate_report.py")


def test_slice_examples_reports_remaining():
    shown, remaining = mod.slice_examples(list(range(20)), 15)
    assert len(shown) == 15 and remaining == 5


def test_slice_examples_no_remaining():
    shown, remaining = mod.slice_examples([1, 2, 3], 15)
    assert shown == [1, 2, 3] and remaining == 0


def test_mapped_badge_text():
    assert "mapped" in mod.mapped_badge(True).lower()
    assert "not" in mod.mapped_badge(False).lower()


def test_competitor_heading_adds_wirkstoff():
    assert mod.competitor_heading("Saxenda", "Liraglutid") == "Saxenda (Liraglutid)"


def test_competitor_heading_no_generic():
    assert mod.competitor_heading("Saxenda", "") == "Saxenda"
    # Avoid redundant '(Saxenda)' when brand == generic.
    assert mod.competitor_heading("Saxenda", "saxenda") == "Saxenda"


def test_cross_competitor_stats_counts_multi():
    claims = [
        {"speaker_name": "Dr. A", "mapped": True, "s_customer_id": "c1",
         "competitor": "Saxenda", "sentiment": "positive"},
        {"speaker_name": "Dr. A", "mapped": True, "s_customer_id": "c1",
         "competitor": "Mounjaro", "sentiment": "negative"},
        {"speaker_name": "Dr. B", "mapped": False, "s_customer_id": "",
         "competitor": "Saxenda", "sentiment": "neutral"},
    ]
    s = mod.cross_competitor_stats(claims)
    assert s["total_doctors"] == 2
    assert s["n_multi"] == 1
    assert s["multi_doctors"][0]["name"] == "Dr. A"
    assert s["multi_doctors"][0]["competitors"] == ["Mounjaro", "Saxenda"]
    assert s["mapped_doctors"] == 1 and s["unmapped_doctors"] == 1
    # Reach: Saxenda has 2 distinct doctors, Mounjaro 1 → Saxenda leads.
    assert s["competitor_reach"][0] == {"competitor": "Saxenda", "generic": "",
                                        "n_doctors": 2}
    assert s["competitor_reach"][1]["competitor"] == "Mounjaro"


SYNTH = {"indication": "Adipositas", "client_drug": "Ozempic",
         "claims": [
             {"speaker_name": "A", "mapped": True, "s_customer_id": "c1", "competitor": "Saxenda",
              "generic": "Liraglutid", "statement": "x", "verbatim_quote": "q",
              "sentiment": "positive", "confidence": "high",
              "citation": {"website_id": "w1", "url": "http://a"}, "verified": True}],
         "competitor_summaries": [{"competitor": "Saxenda", "generic": "Liraglutid",
             "distribution_split": {"all": {"positive": 1, "neutral": 0, "negative": 0, "ambivalent": 0},
                                    "mapped": {"positive": 1, "neutral": 0, "negative": 0, "ambivalent": 0},
                                    "unmapped": {"positive": 0, "neutral": 0, "negative": 0, "ambivalent": 0}},
             "market_view": "mv"}],
         "overall_summary": "os"}


def test_write_excel_one_row_per_claim(tmp_path):
    path = str(tmp_path / "out.xlsx")
    mod.write_excel(SYNTH, path)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["Grounded Claims"]
    assert ws.max_row == 2  # header + 1 claim
    headers = [c.value for c in ws[1]]
    assert "Mapped" in headers and "Verbatim Quote" in headers and "Wirkstoff" in headers


def test_reports_render_without_error():
    a = mod.build_report_a(SYNTH, 15, "2026-07-03 12:00:00")
    b = mod.build_report_b(SYNTH, "2026-07-03 12:00:00")
    assert "Competitor Intelligence Report" in a and "Saxenda (Liraglutid)" in a
    assert "Cross-Competitor Insights" in a and "Per-HCP Drill-Down" not in a
    assert "Most discussed by distinct doctors" in a
    assert "Plain-Language Guide" in b


def test_visible_claims_drops_coi():
    claims = [
        {"speaker_name": "A", "competitor": "Wegovy",
         "verbatim_quote": "Semaglutid senkt das Gewicht deutlich.",
         "statement": "efficacy"},
        {"speaker_name": "B", "competitor": "Wegovy",
         "verbatim_quote": ("Ich erhalte Forschungsgelder von Novo Nordisk. "
                            "Ich halte auch Aktien der Firma."),
         "statement": "receives funding and holds stocks"},
    ]
    out = mod._visible_claims(claims)
    assert len(out) == 1 and out[0]["speaker_name"] == "A"


def test_overall_distribution_sums_across_competitors():
    dist_by_comp = {
        "A": {"positive": 2, "neutral": 1, "negative": 0, "ambivalent": 0},
        "B": {"positive": 3, "neutral": 0, "negative": 1, "ambivalent": 2},
    }
    assert mod.overall_distribution(dist_by_comp) == {
        "positive": 5, "neutral": 1, "negative": 1, "ambivalent": 2}


def test_competitor_distributions_from_claims():
    claims = [
        {"competitor": "Saxenda", "sentiment": "positive"},
        {"competitor": "Saxenda", "sentiment": "negative"},
        {"competitor": "Mounjaro", "sentiment": "neutral"},
    ]
    d = mod.competitor_distributions(claims)
    assert d["Saxenda"] == {"positive": 1, "neutral": 0, "negative": 1, "ambivalent": 0}
    assert d["Mounjaro"]["neutral"] == 1


def test_overview_table_reflects_filtered_not_summaries():
    # summaries say Wegovy neutral=1, but the only visible claim is positive.
    summaries = [{"competitor": "Wegovy", "generic": "Semaglutid",
                  "distribution_split": {"all": {"positive": 1, "neutral": 1,
                                                 "negative": 0, "ambivalent": 0}}}]
    claims = [{"competitor": "Wegovy", "sentiment": "positive"}]
    dist_by_comp = mod.competitor_distributions(claims)
    stats = mod.cross_competitor_stats(claims)
    html = mod._panel_overview(summaries, claims, "summary", stats, dist_by_comp)
    # Wegovy row must be pos1 neu0 neg0 amb0 (from filtered claims), NOT pos1 neu1 (summaries)
    assert "<td>1</td><td>0</td><td>0</td><td>0</td>" in html
    assert "<td>1</td><td>1</td>" not in html


def test_tab_id_slugifies():
    assert mod.tab_id("Saxenda (Liraglutid)") == "tab-saxenda-liraglutid"
    assert mod.tab_id("Insgesamt") == "tab-insgesamt"
    assert mod.tab_id("Most active voices") == "tab-most-active-voices"


def test_report_a_has_sidebar_and_panels():
    a = mod.build_report_a(SYNTH, 15, "2026-07-03 12:00:00")
    # progressive-enhancement scaffolding preserved
    assert "function showTab" in a
    assert "js-tabs" in a
    # two-column sidebar layout markers
    assert 'class="layout"' in a
    assert 'class="sidebar"' in a
    assert 'class="content"' in a
    # grouped, non-clickable section headers, in order
    import re as _re
    labels = _re.findall(r'nav-group-label">([^<]+)<', a)
    assert labels == ["OVERVIEW", "BY COMPETITOR", "ACROSS ALL DRUGS", "ABOUT"]
    # items are nav links with matching panel ids
    assert 'class="nav-item' in a
    assert 'href="#tab-insgesamt"' in a and 'id="tab-insgesamt"' in a
    assert 'href="#tab-saxenda-liraglutid"' in a and 'id="tab-saxenda-liraglutid"' in a
    # all expected item labels present
    for label in ("Insgesamt", "Doctors weighing", "Most active voices", "Methodology"):
        assert label in a
    # overview content + substrings the render test relies on
    assert "Overall sentiment" in a
    assert "Cross-Competitor Insights" in a
    assert "Most discussed by distinct doctors" in a
    # the old flat tab bar is gone
    assert 'class="tabs"' not in a
