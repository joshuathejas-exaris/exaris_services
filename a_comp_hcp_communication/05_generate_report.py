#!/usr/bin/env python3
"""Stage 05 — Render the grounded-claim results into three HTML reports + Excel.

Reads Stage 04's synthesis.json (required). Each grounded claim carries a speaker,
a mapped/unmapped flag (+ S_CUSTOMER_ID when mapped), a verbatim quote, sentiment,
confidence, and a source citation. The HTML reports show 10-15 examples per section
with a "+ N more — see Excel" note; the Excel export holds the FULL result set.

All HTML is fully offline: inline CSS, inline SVG charts, no external assets.

Resume-safe: skipped if a report already exists in results/ unless --force.
"""

import argparse
import configparser
import glob
import html
import json
import logging
import os
import re
import sys
import time
from collections import OrderedDict, defaultdict
from typing import Dict, List, Tuple

# Each stage file adds the repo root to sys.path so it can import from shared/.
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from pipeline_common import is_coi_disclosure  # noqa: E402

_HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(_HERE, "config.ini")
SYNTHESIS_PATH = os.path.join(_HERE, "data", "synthesis.json")
RESULTS_DIR = os.path.join(_HERE, "results")

SENTIMENT_LABELS = ("positive", "neutral", "negative", "ambivalent")
LABEL_COLORS = {
    "positive": "#2e7d32",
    "neutral": "#607d8b",
    "negative": "#c62828",
    "ambivalent": "#f59e0b",
    "no_data": "#cbd5e1",
}
CONFIDENCE_COLORS = {"high": "#2e7d32", "medium": "#f59e0b", "low": "#9ca3af"}

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("stage05")


# --------------------------------------------------------------------------- #
# Config / input
# --------------------------------------------------------------------------- #
def load_config(path: str = CONFIG_PATH) -> configparser.ConfigParser:
    if not os.path.exists(path):
        log.error("Config file not found: %s", path)
        sys.exit(1)
    config = configparser.ConfigParser()
    config.read(path)
    return config


def load_synthesis(path: str = SYNTHESIS_PATH) -> dict:
    if not os.path.exists(path):
        log.error("Input %s not found — run Stage 04 (04_synthesize.py) first.", path)
        sys.exit(1)
    with open(path, encoding="utf-8") as fh:
        data = json.load(fh)
    if not isinstance(data, dict):
        log.error("Expected %s to contain a JSON object.", path)
        sys.exit(1)
    return data


# --------------------------------------------------------------------------- #
# Pure helpers (unit-tested)
# --------------------------------------------------------------------------- #
def slice_examples(items, n):
    """Return (first n items, count of the remainder)."""
    if n is None or n < 0:
        return list(items), 0
    return list(items[:n]), max(0, len(items) - n)


def mapped_badge(mapped: bool) -> str:
    """Coloured pill flagging whether a speaker maps to a known HCP record."""
    if mapped:
        return '<span class="badge" style="background:#1565c0">mapped HCP</span>'
    return '<span class="badge" style="background:#9333ea">not mapped</span>'


def competitor_heading(competitor: str, generic: str) -> str:
    """'Brand (Wirkstoff)' — or just the brand when no generic is known."""
    competitor = (competitor or "").strip()
    generic = (generic or "").strip()
    if generic and generic.lower() != competitor.lower():
        return f"{competitor} ({generic})"
    return competitor


def _hcp_key(claim: dict) -> str:
    """Stable per-doctor key: customer id when mapped, else normalised name."""
    return claim.get("s_customer_id") or f"name::{(claim.get('speaker_name') or '').strip().lower()}"


def cross_competitor_stats(claims: List[dict]) -> dict:
    """Aggregate per-doctor cross-competitor activity from the flat claim list.

    Returns:
      total_doctors      — distinct doctors with at least one grounded statement
      n_multi            — how many of them discuss 2+ distinct competitors
      multi_doctors      — those doctors, richest first: {name, mapped, s_customer_id,
                           n_statements, competitors[], sentiments{label:count}}
      top_voices         — doctors with the most statements (any competitor count)
      mapped_doctors / unmapped_doctors — distinct-doctor counts by mapped flag
      competitor_reach   — per competitor, distinct-doctor count, richest first:
                           {competitor, generic, n_doctors}
    """
    doctors: Dict[str, dict] = {}
    comp_doctors: Dict[str, set] = defaultdict(set)
    comp_generic: Dict[str, str] = {}
    for c in claims:
        key = _hcp_key(c)
        comp = c.get("competitor", "")
        comp_doctors[comp].add(key)
        comp_generic.setdefault(comp, c.get("generic", "") or "")
        d = doctors.setdefault(key, {
            "name": c.get("speaker_name") or "Unknown",
            "mapped": bool(c.get("mapped")),
            "s_customer_id": c.get("s_customer_id") or "",
            "competitors": set(),
            "n_statements": 0,
            "sentiments": defaultdict(int),
        })
        d["competitors"].add(comp)
        d["n_statements"] += 1
        d["sentiments"][c.get("sentiment", "")] += 1
        if c.get("speaker_name") and d["name"] == "Unknown":
            d["name"] = c["speaker_name"]

    competitor_reach = sorted(
        ({"competitor": comp, "generic": comp_generic.get(comp, ""),
          "n_doctors": len(keys)} for comp, keys in comp_doctors.items() if comp),
        key=lambda r: r["n_doctors"], reverse=True)

    packed = []
    for d in doctors.values():
        packed.append({
            "name": d["name"], "mapped": d["mapped"], "s_customer_id": d["s_customer_id"],
            "n_statements": d["n_statements"],
            "competitors": sorted(x for x in d["competitors"] if x),
            "sentiments": dict(d["sentiments"]),
        })
    multi = [d for d in packed if len(d["competitors"]) >= 2]
    multi.sort(key=lambda d: (len(d["competitors"]), d["n_statements"]), reverse=True)
    top_voices = sorted(packed, key=lambda d: d["n_statements"], reverse=True)
    return {
        "total_doctors": len(packed),
        "n_multi": len(multi),
        "multi_doctors": multi,
        "top_voices": top_voices,
        "mapped_doctors": sum(1 for d in packed if d["mapped"]),
        "unmapped_doctors": sum(1 for d in packed if not d["mapped"]),
        "competitor_reach": competitor_reach,
    }


def competitor_distributions(claims: List[dict]) -> Dict[str, Dict[str, int]]:
    """Per-competitor sentiment counts computed from the (already COI-filtered) claims."""
    out: Dict[str, Dict[str, int]] = {}
    for c in claims:
        comp = c.get("competitor", "")
        d = out.setdefault(comp, {s: 0 for s in SENTIMENT_LABELS})
        s = c.get("sentiment")
        if s in d:
            d[s] += 1
    return out


def overall_distribution(dist_by_comp: Dict[str, Dict[str, int]]) -> Dict[str, int]:
    """Aggregate sentiment counts across all competitors' (filtered) distributions."""
    total = {s: 0 for s in SENTIMENT_LABELS}
    for d in dist_by_comp.values():
        for s in SENTIMENT_LABELS:
            total[s] += int(d.get(s, 0) or 0)
    return total


def tab_id(label: str) -> str:
    """Slugify a tab label into a stable DOM id (e.g. 'tab-saxenda-liraglutid')."""
    slug = re.sub(r"[^a-z0-9]+", "-", (label or "").lower()).strip("-")
    return "tab-" + (slug or "x")


# --------------------------------------------------------------------------- #
# HTML helpers
# --------------------------------------------------------------------------- #
def esc(value) -> str:
    return html.escape(str(value if value is not None else ""))


def link(url: str) -> str:
    url = (url or "").strip()
    if not url:
        return ""
    if url.lower().startswith(("http://", "https://")):
        shown = url if len(url) <= 70 else url[:67] + "…"
        return f'<a href="{esc(url)}" target="_blank" rel="noopener noreferrer">{esc(shown)}</a>'
    return esc(url)


def label_badge(label: str) -> str:
    color = LABEL_COLORS.get(label, "#607d8b")
    return f'<span class="badge" style="background:{color}">{esc(label)}</span>'


def confidence_badge(confidence: str) -> str:
    if not confidence:
        return ""
    color = CONFIDENCE_COLORS.get(confidence, "#9ca3af")
    return (f'<span class="badge badge-outline" style="border-color:{color};'
            f'color:{color}">{esc(confidence)}</span>')


def svg_distribution_chart(distribution: Dict[str, int]) -> str:
    counts = [(lbl, int(distribution.get(lbl, 0))) for lbl in SENTIMENT_LABELS]
    total = sum(c for _, c in counts)
    if total == 0:
        return '<p class="muted">No grounded verdicts to chart.</p>'
    max_count = max(c for _, c in counts) or 1
    row_h, bar_h, gap = 34, 20, 14
    label_w, track_w = 96, 320
    width = label_w + track_w + 60
    height = len(counts) * row_h + gap
    rows = []
    for i, (lbl, count) in enumerate(counts):
        y = gap + i * row_h
        bar_w = int((count / max_count) * track_w)
        color = LABEL_COLORS.get(lbl, "#607d8b")
        rows.append(
            f'<text x="{label_w - 8}" y="{y + bar_h - 5}" text-anchor="end" '
            f'class="svg-label">{esc(lbl)}</text>'
            f'<rect x="{label_w}" y="{y}" width="{track_w}" height="{bar_h}" '
            f'rx="3" fill="#eef1f4"/>'
            f'<rect x="{label_w}" y="{y}" width="{bar_w}" height="{bar_h}" '
            f'rx="3" fill="{color}"/>'
            f'<text x="{label_w + bar_w + 8}" y="{y + bar_h - 5}" '
            f'class="svg-count">{count}</text>')
    return (f'<svg viewBox="0 0 {width} {height}" width="100%" '
            f'style="max-width:{width}px" role="img" '
            f'aria-label="Sentiment distribution bar chart">' + "".join(rows) + "</svg>")


BASE_CSS = """
* { box-sizing: border-box; }
body {
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica,
    Arial, sans-serif;
  color: #1f2933; background: #ffffff; margin: 0;
  line-height: 1.55; font-size: 15px;
}
.wrap { max-width: 960px; margin: 0 auto; padding: 32px 24px 64px; }
h1 { font-size: 28px; margin: 0 0 4px; }
h2 { font-size: 21px; margin: 40px 0 12px; padding-bottom: 6px;
     border-bottom: 2px solid #e5e9ee; }
h3 { font-size: 17px; margin: 24px 0 8px; }
p { margin: 8px 0; }
a { color: #1565c0; word-break: break-word; }
.muted { color: #6b7280; }
.subtitle { color: #6b7280; margin: 0 0 8px; }
.meta { color: #9aa5b1; font-size: 13px; margin-top: 4px; }
.card { border: 1px solid #e5e9ee; border-radius: 8px; padding: 16px 18px;
        margin: 16px 0; background: #fcfdfe; }
.badge { display: inline-block; padding: 2px 9px; border-radius: 999px;
         color: #fff; font-size: 12px; font-weight: 600; letter-spacing: .2px; }
.badge-outline { background: transparent !important; border: 1.5px solid; }
.quote { border-left: 3px solid #cbd5e1; padding: 4px 0 4px 14px; margin: 8px 0;
         color: #334155; font-style: italic; }
table { border-collapse: collapse; width: 100%; margin: 12px 0; font-size: 14px; }
th, td { border: 1px solid #e5e9ee; padding: 8px 10px; text-align: left;
         vertical-align: top; }
th { background: #f3f6f9; font-weight: 600; }
.scroll { overflow-x: auto; }
.svg-label { font-size: 12px; fill: #374151; }
.svg-count { font-size: 12px; fill: #374151; font-weight: 600; }
.kpis { display: flex; flex-wrap: wrap; gap: 12px; margin: 16px 0; }
.kpi { flex: 1 1 140px; border: 1px solid #e5e9ee; border-radius: 8px;
       padding: 14px 16px; background: #fcfdfe; }
.kpi .n { font-size: 26px; font-weight: 700; }
.kpi .l { color: #6b7280; font-size: 13px; }
.toc a { display: inline-block; margin-right: 16px; }
.more { color: #6b7280; font-size: 13px; margin: 6px 0 2px; font-style: italic; }
pre { background: #0f172a; color: #e2e8f0; padding: 16px; border-radius: 8px;
      overflow-x: auto; font-size: 13px; line-height: 1.4; }
code { background: #eef1f4; padding: 1px 5px; border-radius: 4px; font-size: 13px; }
pre code { background: transparent; padding: 0; }
footer { margin-top: 48px; padding-top: 16px; border-top: 1px solid #e5e9ee;
         color: #9aa5b1; font-size: 12px; }
ul, ol { margin: 8px 0; padding-left: 22px; }
li { margin: 4px 0; }
.hcp-head { display: flex; align-items: baseline; gap: 10px; flex-wrap: wrap; }
.layout { display: flex; gap: 28px; align-items: flex-start; margin: 24px 0 8px; }
.sidebar { flex: 0 0 220px; position: sticky; top: 16px; align-self: flex-start; }
.content { flex: 1 1 auto; min-width: 0; }
.nav-group-label { text-transform: uppercase; letter-spacing: .6px; font-size: 11px;
                   font-weight: 700; color: #9aa5b1; margin: 18px 0 6px; }
.nav-group-label:first-child { margin-top: 0; }
.nav-item { display: block; padding: 6px 10px; margin: 2px 0; border-radius: 6px;
            color: #374151; font-size: 14px; text-decoration: none;
            border-left: 3px solid transparent; }
.nav-item:hover { background: #f3f6f9; color: #1565c0; }
.nav-item.active { background: #eef4fb; color: #1565c0; font-weight: 600;
                   border-left-color: #1565c0; }
body.js-tabs .panel { display: none; }
body.js-tabs .panel.active { display: block; }
@media (max-width: 720px) {
  .layout { flex-direction: column; gap: 8px; }
  .sidebar { position: static; flex-basis: auto; width: 100%; }
}
"""

TAB_SCRIPT = """
<script>
function showTab(id){
  var ps = document.querySelectorAll('.panel');
  for (var i = 0; i < ps.length; i++){ ps[i].classList.toggle('active', ps[i].id === id); }
  var ns = document.querySelectorAll('.nav-item');
  for (var j = 0; j < ns.length; j++){
    var on = ns[j].getAttribute('href') === '#' + id;
    ns[j].classList.toggle('active', on);
    if (on) { ns[j].setAttribute('aria-current', 'page'); }
    else { ns[j].removeAttribute('aria-current'); }
  }
  return false;
}
document.addEventListener('DOMContentLoaded', function(){
  document.body.classList.add('js-tabs');
  var f = document.querySelector('.panel');
  if (f) { showTab(f.id); }
});
</script>
"""


def _render_sidebar(groups: "List[Tuple[str, List[Tuple[str, str]]]]") -> str:
    """groups: list of (group_header, [(item_label, panel_html), ...]).

    Renders a left nav sidebar (grouped, non-clickable headers) beside a content pane
    of panels. The first item overall is active on load. Degrades to a full-scroll page
    when JS is disabled (no panel is hidden by default CSS). Empty groups are skipped."""
    nav = ['<nav class="sidebar" role="tablist" aria-label="Report sections">']
    panels = []
    first = True
    for group_label, items in groups:
        if not items:
            continue
        nav.append(f'<div class="nav-group-label">{esc(group_label)}</div>')
        for item_label, body in items:
            tid = tab_id(item_label)
            active = " active" if first else ""
            current = ' aria-current="page"' if first else ""
            nav.append(f'<a class="nav-item{active}" role="tab" id="{tid}-btn" '
                       f'href="#{tid}" aria-controls="{tid}"{current} '
                       f'onclick="return showTab(\'{tid}\')">{esc(item_label)}</a>')
            panels.append(f'<section class="panel{active}" role="tabpanel" id="{tid}" '
                          f'aria-labelledby="{tid}-btn">\n{body}\n</section>')
            first = False
    nav.append("</nav>")
    content = '<main class="content">\n' + "\n".join(panels) + "\n</main>"
    return '<div class="layout">\n' + "\n".join(nav) + "\n" + content + "\n</div>"


def html_document(title: str, body: str) -> str:
    return ("<!DOCTYPE html>\n"
            '<html lang="en">\n<head>\n<meta charset="utf-8">\n'
            '<meta name="viewport" content="width=device-width, initial-scale=1">\n'
            f"<title>{esc(title)}</title>\n"
            f"<style>{BASE_CSS}</style>\n</head>\n<body>\n"
            f'<div class="wrap">\n{body}\n</div>\n</body>\n</html>\n')


def footer_html(timestamp: str) -> str:
    return (f"<footer>Generated {esc(timestamp)} · Service 1.2 — Competitive HCP "
            "Communication Monitoring · All data offline, no external assets.</footer>")


# --------------------------------------------------------------------------- #
# Claim grouping
# --------------------------------------------------------------------------- #
def _visible_claims(claims: List[dict]) -> List[dict]:
    """Drop financial COI disclosures so they never reach the report or Excel."""
    out = []
    for c in claims:
        if is_coi_disclosure(c.get("verbatim_quote", ""), c.get("statement", "")):
            log.info("Filtered COI disclosure: %s / %s",
                     c.get("speaker_name"), c.get("competitor"))
            continue
        out.append(c)
    return out


def claims_by_competitor(claims: List[dict]) -> "OrderedDict[str, List[dict]]":
    grouped: "OrderedDict[str, List[dict]]" = OrderedDict()
    for c in claims:
        grouped.setdefault(c.get("competitor", ""), []).append(c)
    return grouped


def claims_by_hcp(claims: List[dict]) -> "OrderedDict[str, List[dict]]":
    """Group by S_CUSTOMER_ID when mapped, else by normalised speaker name."""
    grouped: "OrderedDict[str, List[dict]]" = OrderedDict()
    for c in claims:
        key = c.get("s_customer_id") or f"name::{(c.get('speaker_name') or '').strip().lower()}"
        grouped.setdefault(key, []).append(c)
    return grouped


def _claim_example_html(c: dict) -> str:
    quote = (c.get("verbatim_quote") or "").strip()
    cid = (c.get("s_customer_id") or "").strip()
    cid_html = f' <span class="meta">{esc(cid)}</span>' if cid else ""
    src = link((c.get("citation") or {}).get("url", ""))
    parts = ['<div style="margin:12px 0">',
             '<div class="hcp-head">'
             f"<strong>{esc(c.get('speaker_name') or 'Unknown')}</strong> "
             f"{mapped_badge(bool(c.get('mapped')))}{cid_html} "
             f"{label_badge(c.get('sentiment', ''))} "
             f"{confidence_badge(c.get('confidence', ''))}</div>"]
    if quote:
        parts.append(f'<div class="quote">“{esc(quote)}”</div>')
    if c.get("statement"):
        parts.append(f'<p class="muted">{esc(c["statement"])}</p>')
    if src:
        parts.append(f'<p class="meta">Source: {src}</p>')
    parts.append("</div>")
    return "".join(parts)


# --------------------------------------------------------------------------- #
# Report A — Competitor Intelligence
# --------------------------------------------------------------------------- #
def _overview_sentiment_table(summaries: List[dict],
                              dist_by_comp: Dict[str, Dict[str, int]]) -> str:
    rows = ['<div class="scroll"><table>',
            "<tr><th>Competitor</th><th>positive</th><th>neutral</th>"
            "<th>negative</th><th>ambivalent</th></tr>"]
    for cs in summaries:
        comp = cs.get("competitor", "")
        d = dist_by_comp.get(comp, {})
        rows.append(
            f"<tr><td>{esc(competitor_heading(comp, cs.get('generic', '')))}</td>"
            f"<td>{int(d.get('positive', 0) or 0)}</td>"
            f"<td>{int(d.get('neutral', 0) or 0)}</td>"
            f"<td>{int(d.get('negative', 0) or 0)}</td>"
            f"<td>{int(d.get('ambivalent', 0) or 0)}</td></tr>")
    rows.append("</table></div>")
    return "".join(rows)


def _panel_overview(summaries, claims, overall, stats, dist_by_comp) -> str:
    n_mapped = sum(1 for c in claims if c.get("mapped"))
    n_unmapped = len(claims) - n_mapped
    p = []
    # KPI tiles
    p.append('<div class="kpis">'
             f'<div class="kpi"><div class="n">{len(summaries)}</div>'
             '<div class="l">Competitors</div></div>'
             f'<div class="kpi"><div class="n">{stats["total_doctors"]}</div>'
             '<div class="l">Doctors with a grounded statement</div></div>'
             f'<div class="kpi"><div class="n">{len(claims)}</div>'
             '<div class="l">Grounded statements</div></div>'
             f'<div class="kpi"><div class="n">{n_mapped} / {n_unmapped}</div>'
             '<div class="l">Mapped / not mapped</div></div>'
             "</div>")
    # Overall sentiment overview
    p.append("<h2>Overall sentiment across all competitors</h2>")
    p.append('<div class="card">')
    p.append(svg_distribution_chart(overall_distribution(dist_by_comp)))
    p.append(_overview_sentiment_table(summaries, dist_by_comp))
    p.append("</div>")
    # Executive summary
    p.append("<h2>Executive Summary</h2>")
    p.append(f"<p>{esc(overall)}</p>" if overall
             else '<p class="muted">No overall summary was produced for this run.</p>')
    # Cross-competitor KPIs + reach
    p.append("<h2>Cross-Competitor Insights</h2>")
    p.append('<div class="kpis">'
             f'<div class="kpi"><div class="n">{stats["total_doctors"]}</div>'
             '<div class="l">Distinct doctors</div></div>'
             f'<div class="kpi"><div class="n">{stats["n_multi"]}</div>'
             '<div class="l">Discuss 2+ competitors</div></div>'
             f'<div class="kpi"><div class="n">{stats["mapped_doctors"]}</div>'
             '<div class="l">Mapped HCPs</div></div>'
             f'<div class="kpi"><div class="n">{stats["unmapped_doctors"]}</div>'
             '<div class="l">Not-mapped doctors</div></div>'
             "</div>")
    reach = stats["competitor_reach"]
    p.append("<h3>Reach — distinct doctors per competitor</h3>")
    if not reach:
        p.append('<p class="muted">No competitors had grounded statements this run.</p>')
    else:
        top = reach[0]
        p.append('<p>Most discussed by distinct doctors: '
                 f'<strong>{esc(competitor_heading(top["competitor"], top["generic"]))}</strong> '
                 f'— {top["n_doctors"]} doctor(s).</p>')
        p.append('<div class="scroll"><table>')
        p.append("<tr><th>Competitor</th><th>Distinct doctors</th></tr>")
        for r in reach:
            p.append(f"<tr><td>{esc(competitor_heading(r['competitor'], r['generic']))}</td>"
                     f"<td>{r['n_doctors']}</td></tr>")
        p.append("</table></div>")
    return "\n".join(p)


def _panel_competitor(cs, by_comp, dist_by_comp, examples_per_section) -> str:
    competitor = cs.get("competitor", "")
    dist = dist_by_comp.get(competitor, {})
    market_view = (cs.get("market_view") or "").strip()
    p = ['<div class="card">', svg_distribution_chart(dist)]
    if market_view:
        p.append(f"<p>{esc(market_view)}</p>")
    else:
        p.append('<p class="muted">No market-view narrative available.</p>')
    comp_claims = by_comp.get(competitor, [])
    shown, remaining = slice_examples(comp_claims, examples_per_section)
    if shown:
        p.append("<p><strong>Grounded HCP statements</strong></p>")
        for c in shown:
            p.append(_claim_example_html(c))
        if remaining > 0:
            p.append(f'<p class="more">+ {remaining} more statement(s) for '
                     f"{esc(competitor)} — see the Excel export.</p>")
    else:
        p.append('<p class="muted">No grounded statements for this competitor.</p>')
    p.append("</div>")
    return "\n".join(p)


def _panel_multi(stats, examples_per_section) -> str:
    p = ['<p class="muted">Voices comparing several competitor drugs are the most '
         "commercially interesting — they signal where the market conversation "
         "overlaps.</p>"]
    shown_multi, remaining_multi = slice_examples(stats["multi_doctors"], examples_per_section)
    if not shown_multi:
        p.append('<p class="muted">No doctor discussed more than one competitor in this run.</p>')
        return "\n".join(p)
    p.append('<div class="scroll"><table>')
    p.append("<tr><th>Doctor</th><th></th><th>Competitors</th>"
             "<th># statements</th><th>Sentiment mix</th></tr>")
    for d in shown_multi:
        comps = ", ".join(esc(x) for x in d["competitors"])
        senti = ", ".join(f"{k}:{v}" for k, v in d["sentiments"].items() if k)
        cid = f' <span class="meta">{esc(d["s_customer_id"])}</span>' if d["s_customer_id"] else ""
        p.append(f"<tr><td><strong>{esc(d['name'])}</strong>{cid}</td>"
                 f"<td>{mapped_badge(d['mapped'])}</td>"
                 f"<td>{comps} <span class='meta'>({len(d['competitors'])})</span></td>"
                 f"<td>{d['n_statements']}</td>"
                 f"<td>{esc(senti)}</td></tr>")
    p.append("</table></div>")
    if remaining_multi > 0:
        p.append(f'<p class="more">+ {remaining_multi} more multi-competitor doctor(s) '
                 "— see the Excel export.</p>")
    return "\n".join(p)


def _panel_top_voices(stats) -> str:
    p = ['<p class="muted">Doctors with the most grounded statements this run.</p>']
    top = [d for d in stats["top_voices"] if d["n_statements"] > 1][:10]
    if not top:
        p.append('<p class="muted">No doctor made more than one grounded statement.</p>')
        return "\n".join(p)
    p.append('<div class="scroll"><table>')
    p.append("<tr><th>Doctor</th><th></th><th># statements</th><th>Competitors covered</th></tr>")
    for d in top:
        p.append(f"<tr><td><strong>{esc(d['name'])}</strong></td>"
                 f"<td>{mapped_badge(d['mapped'])}</td>"
                 f"<td>{d['n_statements']}</td>"
                 f"<td>{len(d['competitors'])}</td></tr>")
    p.append("</table></div>")
    return "\n".join(p)


def _panel_methodology() -> str:
    return ("<p>Doctors were included only where a database gate confirmed a genuine "
            "doctor discussing the drug's topic, then an LLM extracted verbatim "
            "statements. Every statement is grounded twice: its quote must be present "
            "verbatim in the source, and an independent verification pass confirms the "
            "named doctor actually expresses that view. Mapped statements resolve to a "
            "known HCP record; unmapped ones are genuine doctor statements from the "
            "same sources without a customer match. Financial conflict-of-interest "
            "disclosures (funding, shares, advisory-board roles, honoraria) are "
            "excluded from this report and the Excel export.</p>")


def build_report_a(synthesis: dict, examples_per_section: int, timestamp: str) -> str:
    client_drug = (synthesis.get("client_drug") or "").strip() or "the client drug"
    indication = (synthesis.get("indication") or "").strip() or "unspecified"
    claims = _visible_claims(synthesis.get("claims", []) or [])
    summaries = synthesis.get("competitor_summaries", []) or []
    overall = (synthesis.get("overall_summary") or "").strip()

    by_comp = claims_by_competitor(claims)
    stats = cross_competitor_stats(claims)
    dist_by_comp = competitor_distributions(claims)

    # Header (stays above the tab bar, always visible)
    head: List[str] = []
    head.append("<h1>Competitor Intelligence Report</h1>")
    head.append(f'<p class="subtitle">Client drug: <strong>{esc(client_drug)}</strong> · '
                f"Indication: <strong>{esc(indication)}</strong></p>")
    head.append(f'<p class="meta">Generated {esc(timestamp)}</p>')
    head.append(f'<p class="muted">Legend: {mapped_badge(True)} a speaker resolved to a '
                f"known HCP customer record · {mapped_badge(False)} a doctor genuinely "
                "quoted in a source but not in our HCP records.</p>")

    # Sidebar-grouped sections: OVERVIEW / BY COMPETITOR / ACROSS ALL DRUGS / ABOUT
    competitor_items = []
    for cs in summaries:
        label = competitor_heading(cs.get("competitor", ""), cs.get("generic", ""))
        competitor_items.append(
            (label, _panel_competitor(cs, by_comp, dist_by_comp, examples_per_section)))
    groups: "List[Tuple[str, List[Tuple[str, str]]]]" = [
        ("OVERVIEW", [("Insgesamt",
                       _panel_overview(summaries, claims, overall, stats, dist_by_comp))]),
        ("BY COMPETITOR", competitor_items),
        ("ACROSS ALL DRUGS", [
            ("Doctors weighing", _panel_multi(stats, examples_per_section)),
            ("Most active voices", _panel_top_voices(stats))]),
        ("ABOUT", [("Methodology", _panel_methodology())]),
    ]

    body = "\n".join(head) + "\n" + _render_sidebar(groups) + "\n" + TAB_SCRIPT + "\n" \
        + footer_html(timestamp)
    return html_document("Competitor Intelligence Report", body)


# --------------------------------------------------------------------------- #
# Report B — Plain-Language Guide
# --------------------------------------------------------------------------- #
def build_report_b(synthesis: dict, timestamp: str) -> str:
    client_drug = (synthesis.get("client_drug") or "").strip() or "your drug"
    p: List[str] = []
    p.append("<h1>Plain-Language Guide to This Report</h1>")
    p.append('<p class="subtitle">A jargon-free companion to the Competitor '
             "Intelligence Report.</p>")
    p.append(f'<p class="meta">Generated {esc(timestamp)}</p>')

    p.append("<h2>What is this report?</h2>")
    p.append("<p>This report tells you which doctors (healthcare professionals, or HCPs) "
             f"have been talking about the drugs that compete with {esc(client_drug)}, and "
             "whether what they said was broadly positive, negative, or mixed. Every "
             "opinion links back to the exact document it came from.</p>")

    p.append("<h2>How was the data collected?</h2>")
    p.append("<ol>"
             "<li><strong>Find the right documents and doctors.</strong> We start from "
             "records confirming a real doctor is connected to a document that discusses "
             "the competitor drug.</li>"
             "<li><strong>Pull the whole document.</strong> We read the entire source, not "
             "just a snippet, so context is never lost.</li>"
             "<li><strong>Extract only genuine statements.</strong> We keep a statement "
             "only when the named doctor actually says something about the drug — never "
             "when a doctor is simply mentioned on the same page.</li>"
             "<li><strong>Double-check every statement.</strong> The exact quote must "
             "appear in the source, and a second check confirms the doctor really said "
             "it about that drug.</li>"
             "</ol>")

    p.append("<h2>Mapped vs general (not-mapped) doctors</h2>")
    p.append("<p>Some doctors match a record we already hold — we call these "
             "<strong>mapped HCPs</strong>. Others are real doctors quoted in the same "
             "sources whom we could not match to a record — we call these "
             "<strong>not-mapped</strong> (general) doctors. Both are shown, each clearly "
             "flagged, so you can see the wider medical conversation as well as your "
             "known contacts.</p>")

    p.append("<h2>How to read the sentiment labels</h2>")
    p.append("<table>"
             "<tr><th>Label</th><th>What it means</th><th>Example</th></tr>"
             f"<tr><td>{label_badge('positive')}</td>"
             "<td>The doctor spoke favourably about the competitor drug.</td>"
             "<td>“In my practice this medicine has delivered excellent results.”</td></tr>"
             f"<tr><td>{label_badge('neutral')}</td>"
             "<td>Mentioned factually, without praise or criticism.</td>"
             "<td>“This drug is dosed once weekly by injection.”</td></tr>"
             f"<tr><td>{label_badge('negative')}</td>"
             "<td>Spoke critically or unfavourably about it.</td>"
             "<td>“I stopped prescribing it because of the side-effects.”</td></tr>"
             f"<tr><td>{label_badge('ambivalent')}</td>"
             "<td>Both good and bad — a genuinely mixed view.</td>"
             "<td>“It works well, but the cost is hard to justify.”</td></tr>"
             "</table>")

    p.append("<h2>What does confidence mean?</h2>")
    p.append("<ul>"
             f"<li>{confidence_badge('high')} — direct and unambiguous.</li>"
             f"<li>{confidence_badge('medium')} — reasonably clear, some interpretation.</li>"
             f"<li>{confidence_badge('low')} — brief or vague; treat as a soft signal.</li>"
             "</ul>")

    p.append("<h2>What this report does NOT tell you</h2>")
    p.append("<ul>"
             "<li>It does not measure prescribing volume — only what was said.</li>"
             "<li>It only covers statements written in a traceable source.</li>"
             "<li>A doctor not appearing here means no confirmed statement was found — "
             "not that they have no opinion.</li>"
             "<li>Sentiment is a judgement of tone, not clinical effectiveness.</li>"
             "</ul>")

    p.append("<h2>Glossary</h2>")
    p.append("<table>"
             "<tr><th>Term</th><th>Meaning</th></tr>"
             "<tr><td>HCP</td><td>Healthcare professional — a doctor or clinician.</td></tr>"
             "<tr><td>Mapped HCP</td><td>A doctor matched to a record we already hold.</td></tr>"
             "<tr><td>Not mapped</td><td>A real doctor quoted in a source with no record match.</td></tr>"
             "<tr><td>Competitor drug</td><td>A different medicine for the same condition.</td></tr>"
             "<tr><td>Verbatim quote</td><td>The doctor's exact words, copied from the source.</td></tr>"
             "<tr><td>Source document</td><td>The original webpage or publication.</td></tr>"
             "</table>")
    p.append(footer_html(timestamp))
    return html_document("Plain-Language Guide", "\n".join(p))


# --------------------------------------------------------------------------- #
# Report C — Technical Documentation
# --------------------------------------------------------------------------- #
FLOWCHART = """01_identify_competitors.py   --client-drug X [--cf-data|--from-snowflake]
        |
        v  data/competitors.json   { indication, client_drug, competitors[]{brand,generic,source} }
02_retrieve_sources.py       Track A: LLM_VALIDATION gate + scoped vector search + full CONTENT
                             Track B: global vector search (no LLM_VALIDATION)
        |
        v  data/raw_sources.json   [ { competitor, track, mapped_hcps[], sources[]{full_text} } ]
03_wiki_build.py             ingest grounded claims -> deterministic quote-grounding -> LLM verify
        |                    -> map speaker to S_CUSTOMER_ID; write wiki/<ts>/{raw,wiki,schema}
        v  data/knowledge_graph.json   { competitors[]{nodes, claims[]} }
04_synthesize.py             mapped/unmapped sentiment split + market view + overall summary
        |
        v  data/synthesis.json   { claims[], competitor_summaries[], overall_summary }
05_generate_report.py        render HTML x3 + full Excel
        |
        v  results/report_<ts>.html  guide_<ts>.html  technical_<ts>.html  report_<ts>.xlsx"""

SNOWFLAKE_TABLES = [
    ("LLM_VALIDATION", "schema_final", "Layer-1 gate + full document CONTENT + keywords",
     "NEAR_BY, IS_OLD, IS_DOCTOR, COL_KEYWORDS_ORIG, COL_KEYWORDS_EN, CONTENT, WEBSITE_ID, S_CUSTOMER_ID"),
    ("WEBSITES_VERTICAL_CONTENT_FRAME_SINGLE_TBL", "schema_final",
     "Vertical CF mapping (website_id ↔ url)", "WEBSITE_ID, S_CUSTOMER_ID, URL"),
    ("WEBSITES_VERTICAL_EMBEDDINGS_512", "schema_final",
     "Vertical site chunks + 768-dim embeddings", "CHUNK, EMBEDDINGS, WEBSITE_ID"),
    ("WEBSITES_CONTENT_FRAME_SINGLE", "schema_final", "Public website CF mapping",
     "WEBSITE_ID, DOMAIN_VALUE"),
    ("WEBSITES_EMBEDDINGS_512", "schema_final", "Public website chunks + embeddings",
     "CHUNK, EMBEDDINGS, WEBSITE_ID"),
    ("PUBMED_EMBEDDINGS_512", "schema_final", "PubMed chunks + embeddings (Track B)",
     "CHUNK, EMBEDDINGS, WEBSITE_ID"),
    ("CUSTOMER_SOURCE", "schema_tmp", "HCP master record (mapped-HCP roster)",
     "S_FIRSTNAME, S_LASTNAME, S_CITY"),
    ("CONTENT_FRAME_SPEC", "schema_tmp", "CF terms for Stage 01",
     "DE_TERM_1, EN_TERM_1"),
]

CONFIG_REFERENCE = [
    ("[llm_validation] near_by/is_old/is_doctor", "int", "1/0/1", "Revised Layer-1 gate (IN_RELATION dropped)"),
    ("[retrieval] top_chunks_per_wirkstoff", "int", "100", "Best chunks kept per wirkstoff query set"),
    ("[retrieval] min_similarity", "float", "0.65", "Min cosine similarity to keep a chunk"),
    ("[retrieval] max_sources_per_competitor", "int", "40", "Cap on documents fed to the wiki"),
    ("[wiki] ingest_model_id", "str", "eu.amazon.nova-pro-v1:0", "Claim extraction model"),
    ("[wiki] verify_model_id", "str", "qwen.qwen3-235b-…", "Adversarial verification model"),
    ("[wiki] max_source_chars", "int", "24000", "Per-document truncation before ingest"),
    ("[wiki] content_source", "str", "llm_validation", "Full-text source (else chunk_concat)"),
    ("[report] examples_per_section", "int", "15", "Examples shown per HTML section (rest in Excel)"),
]


def build_report_c(synthesis: dict, config: configparser.ConfigParser, timestamp: str) -> str:
    p: List[str] = []
    p.append("<h1>Technical Documentation</h1>")
    p.append('<p class="subtitle">Pipeline internals for engineers — Service 1.2 '
             "Competitive HCP Communication Monitoring.</p>")
    p.append(f'<p class="meta">Generated {esc(timestamp)}</p>')
    p.append('<p class="toc">'
             '<a href="#overview">Overview</a><a href="#tables">Snowflake tables</a>'
             '<a href="#grounding">Grounding model</a><a href="#repro">Reproducing a run</a>'
             '<a href="#config">Config reference</a></p>')

    p.append('<h2 id="overview">1. Pipeline Overview</h2>')
    p.append(f"<pre><code>{esc(FLOWCHART)}</code></pre>")

    p.append('<h2 id="tables">2. Snowflake Tables Used</h2>')
    p.append('<div class="scroll"><table>')
    p.append("<tr><th>Table</th><th>Schema</th><th>Purpose</th><th>Key columns</th></tr>")
    for name, schema, purpose, cols in SNOWFLAKE_TABLES:
        p.append(f"<tr><td><code>{esc(name)}</code></td><td>{esc(schema)}</td>"
                 f"<td>{esc(purpose)}</td><td><code>{esc(cols)}</code></td></tr>")
    p.append("</table></div>")

    p.append('<h2 id="grounding">3. Grounding Model</h2>')
    p.append("<div class='card'><h3>Layer 1 — Revised LLM_VALIDATION gate</h3>"
             "<p><code>NEAR_BY=1 AND IS_OLD=0 AND IS_DOCTOR=1</code> plus the wirkstoff/brand "
             "present in <code>COL_KEYWORDS_ORIG/EN</code>. <code>IN_RELATION</code> is "
             "intentionally NOT gated — it is non-indicative. Yields relevant website IDs "
             "and the mapped-HCP roster.</p></div>")
    p.append("<div class='card'><h3>Layer 2 — Scoped vector search + full content</h3>"
             "<p>VECTOR_COSINE_SIMILARITY restricted to the Layer-1 website IDs picks the "
             "best chunks; the whole document CONTENT is then assembled as the wiki raw "
             "source.</p></div>")
    p.append("<div class='card'><h3>Layer 3 — LLM-wiki: ingest, ground, verify</h3>"
             "<p>Ingest extracts a claim only when a named doctor expresses a view about "
             "the drug. A deterministic check drops any claim whose verbatim quote is not "
             "literally in the source; an adversarial LLM pass then re-confirms speaker + "
             "attribution. <strong>Failure mode prevented:</strong> a doctor named on a "
             "page while the drug sits elsewhere — no genuine statement.</p></div>")

    p.append('<h2 id="repro">4. Reproducing a Run</h2>')
    repro = ("python 01_identify_competitors.py --client-drug \"Ozempic\" --from-snowflake\n"
             "python 02_retrieve_sources.py\n"
             "python 03_wiki_build.py\n"
             "python 04_synthesize.py\n"
             "python 05_generate_report.py\n\n"
             "# Force any stage to rebuild its output:\n"
             "python 05_generate_report.py --force")
    p.append(f"<pre><code>{esc(repro)}</code></pre>")

    p.append('<h2 id="config">5. Config Reference (key params)</h2>')
    p.append('<div class="scroll"><table>')
    p.append("<tr><th>Key</th><th>Type</th><th>Default</th><th>Effect</th></tr>")
    for key, typ, default, effect in CONFIG_REFERENCE:
        p.append(f"<tr><td><code>{esc(key)}</code></td><td>{esc(typ)}</td>"
                 f"<td><code>{esc(default)}</code></td><td>{esc(effect)}</td></tr>")
    p.append("</table></div>")
    p.append(footer_html(timestamp))
    return html_document("Technical Documentation", "\n".join(p))


# --------------------------------------------------------------------------- #
# Excel — full results
# --------------------------------------------------------------------------- #
def write_excel(synthesis, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Grounded Claims"
    headers = ["HCP/Speaker", "Mapped", "S_CUSTOMER_ID", "Competitor", "Wirkstoff",
               "Sentiment", "Confidence", "Statement", "Verbatim Quote", "Source URL",
               "Verified"]
    ws.append(headers)
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="37474F")
    for i, _ in enumerate(headers, 1):
        ws.cell(row=1, column=i).font = hf
        ws.cell(row=1, column=i).fill = fill
    for c in _visible_claims(synthesis.get("claims", []) or []):
        cit = c.get("citation", {}) or {}
        ws.append([c.get("speaker_name", ""), "yes" if c.get("mapped") else "no",
                   c.get("s_customer_id", ""), c.get("competitor", ""), c.get("generic", ""),
                   c.get("sentiment", ""), c.get("confidence", ""), c.get("statement", ""),
                   c.get("verbatim_quote", ""), cit.get("url", ""),
                   "yes" if c.get("verified") else "no"])
    for i, w in enumerate([22, 8, 16, 16, 16, 12, 12, 40, 60, 40, 9], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wrap = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Competitor Summary")
    ws2.append(["Competitor", "Scope", "Positive", "Neutral", "Negative", "Ambivalent",
                "Market View"])
    for i in range(1, 8):
        ws2.cell(row=1, column=i).font = hf
        ws2.cell(row=1, column=i).fill = fill
    for s in synthesis.get("competitor_summaries", []):
        ds = s.get("distribution_split", {})
        for scope in ("all", "mapped", "unmapped"):
            d = ds.get(scope, {})
            ws2.append([s.get("competitor", ""), scope, d.get("positive", 0),
                        d.get("neutral", 0), d.get("negative", 0), d.get("ambivalent", 0),
                        s.get("market_view", "") if scope == "all" else ""])
    for i, w in enumerate([18, 10, 10, 10, 10, 12, 70], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for row in ws2.iter_rows(min_row=2):
        row[-1].alignment = wrap
    wb.save(path)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Stage 05 — render HTML reports + Excel from the synthesis output.")
    parser.add_argument("--force", action="store_true",
                        help="Regenerate even if reports already exist in results/.")
    return parser.parse_args()


def existing_reports() -> List[str]:
    return sorted(glob.glob(os.path.join(RESULTS_DIR, "report_*.html")))


def main() -> None:
    args = parse_args()
    prior = existing_reports()
    if prior and not args.force:
        log.info("%d report(s) already in results/ (e.g. %s) — skipping (use --force).",
                 len(prior), os.path.basename(prior[-1]))
        return

    config = load_config()
    synthesis = load_synthesis()
    examples_per_section = config["report"].getint("examples_per_section", fallback=15)

    claims = synthesis.get("claims", []) or []
    log.info("Loaded %d claim(s), %d competitor summar(y/ies).",
             len(claims), len(synthesis.get("competitor_summaries", []) or []))

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    human_ts = time.strftime("%Y-%m-%d %H:%M:%S")
    os.makedirs(RESULTS_DIR, exist_ok=True)
    report_a_path = os.path.join(RESULTS_DIR, f"report_{timestamp}.html")
    guide_path = os.path.join(RESULTS_DIR, f"guide_{timestamp}.html")
    technical_path = os.path.join(RESULTS_DIR, f"technical_{timestamp}.html")
    excel_path = os.path.join(RESULTS_DIR, f"report_{timestamp}.xlsx")

    with open(report_a_path, "w", encoding="utf-8") as fh:
        fh.write(build_report_a(synthesis, examples_per_section, human_ts))
    log.info("Wrote %s", report_a_path)
    with open(guide_path, "w", encoding="utf-8") as fh:
        fh.write(build_report_b(synthesis, human_ts))
    log.info("Wrote %s", guide_path)
    with open(technical_path, "w", encoding="utf-8") as fh:
        fh.write(build_report_c(synthesis, config, human_ts))
    log.info("Wrote %s", technical_path)
    write_excel(synthesis, excel_path)
    log.info("Wrote %s", excel_path)
    log.info("Stage 05 complete — 3 HTML reports + 1 Excel in results/.")


if __name__ == "__main__":
    main()
